from urllib.parse import unquote
import json
from django.views.decorators.csrf import csrf_exempt
from django.http import HttpResponseForbidden, HttpResponseNotFound, HttpResponse, HttpResponseRedirect
from django.shortcuts import redirect, render, get_object_or_404
from django.contrib.auth.decorators import login_required
from .models import *
from .forms import *
from django.http import JsonResponse
from django.core.paginator import Paginator
from datetime import datetime, timedelta, date, time
from django.db.models import Count, Min, Max, Manager
import random
from django.contrib import messages
from django.forms.models import model_to_dict
from django.utils import timezone
import openpyxl
from openpyxl.styles import Font
from django.apps import apps
from django.core.files.storage import FileSystemStorage
from django.conf import settings
from decimal import Decimal, DivisionUndefined
from django.views.decorators.csrf import csrf_exempt
import requests
from django.db.models import Q

def get_exchange_rates(request):
    today = date.today()
    # Bugün için ExchangeRate var mı kontrol et
    if ExchangeRate.objects.filter(created_at__date=today).exists():
        print("Bugün için zaten döviz kuru verisi mevcut.")
        exchange_rate = ExchangeRate.objects.get(created_at__date=today)
        data = {
            'usd_to_try': exchange_rate.usd_to_try,
            'usd_to_eur': exchange_rate.usd_to_eur,
            'usd_to_rmb': exchange_rate.usd_to_rmb
        }
        return JsonResponse(data)
    else:
        return exchange_rates()

def exchange_rates():
    today = date.today()
    if not ExchangeRate.objects.filter(created_at__date=today).exists():
        url = "https://api.exchangerate-api.com/v4/latest/USD"
        response = requests.get(url)
        if response.status_code == 200:
            data = response.json()
            usd_to_try = data['rates']['TRY']
            usd_to_eur = data['rates']['EUR']
            usd_to_rmb = data['rates']['CNY']

            # Döviz kurlarını kaydet
            exchange_rate = ExchangeRate.objects.create(
                usd_to_try=usd_to_try,
                usd_to_eur=usd_to_eur,
                usd_to_rmb=usd_to_rmb,
                created_at=timezone.now()
            )

            print("Döviz kurları başarıyla kaydedildi.")
            data = {
                'usd_to_try': usd_to_try,
                'usd_to_eur': usd_to_eur,
                'usd_to_rmb': usd_to_rmb
            }
            return JsonResponse(data)
        else:
            print("Veriler çekilemedi. Lütfen daha sonra tekrar deneyin.")
            return JsonResponse({'error': 'Veriler çekilemedi'}, status=500)
    else:
        exchange_rate = ExchangeRate.objects.get(created_at__date=today)
        data = {
            'usd_to_try': exchange_rate.usd_to_try,
            'usd_to_eur': exchange_rate.usd_to_eur,
            'usd_to_rmb': exchange_rate.usd_to_rmb
        }
        return JsonResponse(data)

def can_access_personel(user):
    # Kullanıcının Personel modeline erişim yetkisini kontrol et
    personel = user.personel.first()
    return personel and personel.job in ['Sistem Geliştiricisi', 'Yönetim']

def can_access_personel2(user):
    # Kullanıcının Personel modeline erişim yetkisini kontrol et
    personel = user.personel.first()
    return personel and personel.job in ['Sistem Geliştiricisi', 'Yönetim', 'Operasyon Şefi']

def can_access_personel3(user):
    # Kullanıcının Personel modeline erişim yetkisini kontrol et
    personel = user.personel.first()
    return personel and personel.job in ['Sistem Geliştiricisi', 'Yönetim', 'Muhasebe']



def check_username(request):
    username = request.GET.get('username', '')
    is_taken = User.objects.filter(username=username).exists()
    return JsonResponse({'is_taken': is_taken})

@login_required
def generic_view(request, model):
    if model == "Personel" and not can_access_personel3(request.user):
        return HttpResponseForbidden("Bu bölüme erişim yetkiniz bulunmamaktadır.")
    if model == "Cost" and not can_access_personel2(request.user):
        return HttpResponseForbidden("Bu bölüme erişim yetkiniz bulunmamaktadır.")
    if model == "Activitycost" and not can_access_personel2(request.user):
        return HttpResponseForbidden("Bu bölüme erişim yetkiniz bulunmamaktadır.")

    requestPersonel = Personel.objects.get(user=request.user)
    sirket = requestPersonel.company
    ModelClass = globals()[model]
    model_plural_name = model.capitalize() + 's'
    model_lower_plural_name = model.lower() + 's'
    queryset = ModelClass.objects.filter(company=sirket, is_delete=False)
    form = globals()[f"{model}Form"]()
    fields = ModelClass._meta.fields
    objects_data = []
    for obj in queryset:
        obj_data = {}
        for field in fields:
            obj_data[field.name] = getattr(obj, field.name)
        objects_data.append(obj_data)


    context = {
        'title': model.capitalize(),
        'createtitle': f'CREATE {model.upper()}',
        'listTitle': f'{model.upper()} LIST',
        'createUrl': f'create_{model.lower()}',
        'deleteUrl': f'delete_{model}',
        'form': form,
        'objects_data': objects_data,  # Model nesnelerinin alan değerlerini içeren liste
        'models': queryset,  # Model nesneleri
        'fields': [field.name for field in fields],
        'model_name' : model  # Alan başlıkları
    }

    return render(request, f'tour/pages/generic_main.html', context)

@login_required
def generic_create_view(request, model):
    request_personnel = Personel.objects.get(user=request.user)
    company = request_personnel.company
    model_class = globals().get(model.capitalize())
    if model_class is None:
        return HttpResponseNotFound(f"{model.capitalize()} model not found")

    form_class = globals().get(f"{model.capitalize()}Form")
    if form_class is None:
        return HttpResponseNotFound(f"{model.capitalize()}Form not found")

    if request.method == "POST":
        form = form_class(request.POST or None)
        if form.is_valid():
            print('##############################################')
            print(form)
            print('##############################################')
            new_object = form.save(commit=False)
            new_object.company = company
            if model == "Notification":
                new_object.sender = request_personnel
            new_object.save()

            action_description = f"{model.capitalize()} oluşturdu. ID: {new_object.id}"
            log = UserActivityLog(
                company=company,
                staff=request_personnel,
                action=action_description
            )
            log.save()


            obj_data = {'id': new_object.id}
            for field in model_class._meta.fields:
                obj_data[field.name] = getattr(new_object, field.name)
            context = {
                'fields': [field.name for field in model_class._meta.fields],  # alan adlarını dahil et
                'obj_data': obj_data,
                'model_name': model
            }
            return render(request, 'tour/partials/generic-list.html', context)
        else:
            # Hata log kaydı
            errors = form.errors.as_json()
            action_description = f"{model.capitalize()} Kayıt Oluşturulamadı. Hata: " + errors
            log = UserActivityLog(
                company=company,
                staff=request_personnel,
                action=action_description
            )
            log.save()
    else:
        form = form_class()
    context = {
        'form': form,
        'model_name': model
    }
    return render(request, f'tour/partials/generic-form.html', context)

@login_required
def generic_edit_view(request, model, obj_id):
    # Model sınıfını dinamik olarak yükle
    ModelClass = globals().get(model.capitalize())
    if ModelClass is None:
        return HttpResponseNotFound(f"{model.capitalize()} model not found.")

    # Model objesini getir veya 404 döndür
    obj = get_object_or_404(ModelClass, id=obj_id)
    request_personnel = Personel.objects.get(user=request.user)
    company = request_personnel.company

    # Form sınıfını dinamik olarak yükle
    FormClass = globals().get(f"{model.capitalize()}Form")
    if FormClass is None:
        return HttpResponseNotFound(f"{model.capitalize()}Form not found.")

    if request.method == "POST":
        original_data = {field.name: getattr(obj, field.name) for field in ModelClass._meta.fields}
        form = FormClass(request.POST, instance=obj)
        if form.is_valid():
            updated_obj = form.save()  # Güncellenen objeyi döndürür
            updated_data = {field.name: getattr(updated_obj, field.name) for field in ModelClass._meta.fields}
            changes = [
                f"{key}: {original_data[key]} den {updated_data[key]} ye değişti"
                for key in original_data if original_data[key] != updated_data[key]
            ]

            if changes:
                action_description = f"Güncellendi. {model.capitalize()} {obj_id}: " + ", ".join(changes)
                log = UserActivityLog(
                    company=company,
                    staff=request_personnel,
                    action=action_description
                )
                log.save()
            obj_data = {'id': obj_id}  # ID bilgisini ekleyin
            for field in ModelClass._meta.fields:
                obj_data[field.name] = getattr(obj, field.name)
            context = {
                'form': form,
                'obj_data': obj_data,  # Güncellenmiş obje verileri
                'model_name': model,
                'fields': [field.name for field in ModelClass._meta.fields],
            }
            # Güncellenmiş verilerle liste şablonunu yeniden render et
            return render(request, 'tour/partials/generic-list.html', context)
    else:
        form = FormClass(instance=obj)
        context = {
            'form': form,
            'obj_data': obj,  # Form başlangıcında objeyi gönder
            'model_name': model,
            'fields': [field.name for field in ModelClass._meta.fields],
        }

    return render(request, 'tour/partials/generic-edit-form.html', context)


@login_required
def generic_delete_view(request, model, obj_id):
    ModelClass = globals().get(model.capitalize())
    if ModelClass is None:
        return HttpResponseNotFound(f"{model.capitalize()} model not found.")

    # Model objesini getir veya 404 döndür
    obj = get_object_or_404(ModelClass, id=obj_id)
    request_personnel = Personel.objects.get(user=request.user)
    company = request_personnel.company

    if request.method == "DELETE":
        obj_data = {field.name: getattr(obj, field.name) for field in ModelClass._meta.fields}
        transfer = get_object_or_404(ModelClass, id=obj_id)
        try:
            transfer.is_delete = True
            transfer.save()
            action_description = f"Silindi {model.capitalize()} {obj_id}: " + ", ".join([f"{key}: {value}" for key, value in obj_data.items()])
            log = UserActivityLog(
                company=company,
                staff=request_personnel,
                action=action_description
            )
            log.save()
            return HttpResponse('')
        except Personel.DoesNotExist:
            pass  # veya uygun bir hata mesajı göster
          # Boş bir yanıt döndür

@login_required
def generic_cancel_view(request, model, obj_id):
    ModelClass = globals().get(model.capitalize())
    if ModelClass is None:
        return HttpResponseNotFound(f"{model.capitalize()} model not found.")
    # Model objesini getir veya 404 döndür
    obj = get_object_or_404(ModelClass, id=obj_id)
    # İlgili firma nesnesini al
    obj_data = {'id': obj_id}  # ID bilgisini ekleyin
    for field in ModelClass._meta.fields:
        obj_data[field.name] = getattr(obj, field.name)

    # Firma bilgilerini içeren bir template döndür
    return render(request, 'tour/partials/generic-list.html', {'obj_data': obj_data, 'fields': [field.name for field in ModelClass._meta.fields], 'model_name': model,})

@login_required
def generic_excel_download(request, model):
    # Model sınıfını dinamik olarak yükle
    ModelClass = apps.get_model(app_label='Core', model_name=model)
    if ModelClass is None:
        return HttpResponseNotFound(f"{model} model not found.")

    # Excel dosyasını oluştur
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{model} Template"

    # Başlıkları yaz
    fields = [field for field in ModelClass._meta.fields if field.name not in ['id', 'company', 'is_delete']]
    field_names = [field.verbose_name for field in fields]

    for col_num, field_name in enumerate(field_names, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = field_name
        cell.font = Font(bold=True)

    # HttpResponse ile Excel dosyasını döndür
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={model}_template.xlsx'
    wb.save(response)

    return response

@login_required
def generic_excel_full_download(request, model):
    # Model sınıfını dinamik olarak yükle
    ModelClass = apps.get_model(app_label='Core', model_name=model)
    if ModelClass is None:
        return HttpResponseNotFound(f"{model} model not found.")

    # Excel dosyasını oluştur
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{model} Template"

    # Başlıkları yaz
    fields = [field for field in ModelClass._meta.fields if field.name not in ['id', 'company', 'is_delete']]
    field_names = [field.verbose_name for field in fields]

    for col_num, field_name in enumerate(field_names, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = field_name
        cell.font = Font(bold=True)

    # Verileri yaz
    for row_num, obj in enumerate(ModelClass.objects.filter(is_delete=False), 2):
        for col_num, field in enumerate(fields, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell_value = getattr(obj, field.name)
            cell.value = str(cell_value)  # Tüm veri türlerini stringe çevirerek yaz

    # HttpResponse ile Excel dosyasını döndür
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={model}_full.xlsx'
    wb.save(response)

    return response

@login_required
def generic_excel_upload(request, model):
    print("Request method:", request.method)
    request_Personel = Personel.objects.get(user=request.user)
    company = request_Personel.company
    if request.method == 'POST':
        print("POST request received")
        if 'excel_file' in request.FILES:
            print("Excel file found in request.FILES")
            excel_file = request.FILES['excel_file']
            fs = FileSystemStorage()
            filename = fs.save(excel_file.name, excel_file)
            uploaded_file_url = fs.url(filename)
            print("File saved as:", filename)
            print("Uploaded file URL:", uploaded_file_url)

            # Excel dosyasını yükle
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active

            # Model sınıfını dinamik olarak yükle
            ModelClass = apps.get_model(app_label='Core', model_name=model)
            if ModelClass is None:
                print("Model not found:", model)
                return HttpResponseNotFound(f"{model} model not found.")

            # Alanları alın, id ve company hariç
            fields = [field for field in ModelClass._meta.fields if field.name not in ['id', 'company', 'is_delete']]
            field_names = [field.name for field in fields]
            print("Field names:", field_names)

            # Verileri işleyin ve kaydedin
            for row in ws.iter_rows(min_row=2, values_only=True):  # İlk satır başlık olduğu için 2. satırdan itibaren başla
                obj_data = {field_name: value for field_name, value in zip(field_names, row)}
                print("Object data to save:", obj_data)
                new_object = ModelClass(**obj_data)
                new_object.company = Personel.objects.get(user=request.user).company  # Şirket bilgilerini ekleyin
                new_object.save()
                print("Object saved:", new_object)
                action_description = f"Excel Upload ile {model.capitalize()} oluşturdu. ID: {new_object.id}"
                log = UserActivityLog(
                    company=Personel.objects.get(user=request.user).company,
                    staff=request_Personel,
                    action=action_description
                )
                log.save()

            return redirect("generic_view", model)
        else:
            print("No excel_file in request.FILES")

    return render(request, 'your_app/upload_form.html', {'model': model})


@login_required
def generic_mobile_create_view(request, model):
    request_personnel = Personel.objects.get(user=request.user)
    company = request_personnel.company
    model_class = globals().get(model.capitalize())
    if model_class is None:
        return HttpResponseNotFound(f"{model.capitalize()} model not found")

    form_class = globals().get(f"{model.capitalize()}Form")
    if form_class is None:
        return HttpResponseNotFound(f"{model.capitalize()}Form not found")

    if request.method == "POST":
        form = form_class(request.POST or None)
        if form.is_valid():
            new_object = form.save(commit=False)
            new_object.company = company
            new_object.save()

            action_description = f"{model.capitalize()} oluşturdu. ID: {new_object.id}"
            log = UserActivityLog(
                company=company,
                staff=request_personnel,
                action=action_description
            )
            log.save()


            obj_data = {'id': new_object.id}
            for field in model_class._meta.fields:
                obj_data[field.name] = getattr(new_object, field.name)
            context = {
                'fields': [field.name for field in model_class._meta.fields],  # alan adlarını dahil et
                'obj_data': obj_data,
                'model_name': model
            }
            return render(request, 'tour/partials/mobile-card-card.html', context)
        else:
            # Hata log kaydı
            errors = form.errors.as_json()
            action_description = f"{model.capitalize()} Kayıt Oluşturulamadı. Hata: " + errors
            log = UserActivityLog(
                company=company,
                staff=request_personnel,
                action=action_description
            )
            log.save()
    else:
        form = form_class()
    context = {
        'form': form,
        'model_name': model
    }
    return render(request, f'tour/partials/generic-mobile-form.html', context)


@login_required
def generic_mobile_edit_view(request, model, obj_id):
    # Model sınıfını dinamik olarak yükle
    ModelClass = globals().get(model.capitalize())
    if ModelClass is None:
        return HttpResponseNotFound(f"{model.capitalize()} model not found.")

    # Model objesini getir veya 404 döndür
    obj = get_object_or_404(ModelClass, id=obj_id)
    request_personnel = Personel.objects.get(user=request.user)
    company = request_personnel.company

    # Form sınıfını dinamik olarak yükle
    FormClass = globals().get(f"{model.capitalize()}Form")
    if FormClass is None:
        return HttpResponseNotFound(f"{model.capitalize()}Form not found.")

    if request.method == "POST":
        original_data = {field.name: getattr(obj, field.name) for field in ModelClass._meta.fields}
        form = FormClass(request.POST, instance=obj)
        if form.is_valid():
            updated_obj = form.save()  # Güncellenen objeyi döndürür
            updated_data = {field.name: getattr(updated_obj, field.name) for field in ModelClass._meta.fields}
            changes = [
                f"{key}: {original_data[key]} den {updated_data[key]} ye değişti"
                for key in original_data if original_data[key] != updated_data[key]
            ]

            if changes:
                action_description = f"Güncellendi. {model.capitalize()} {obj_id}: " + ", ".join(changes)
                log = UserActivityLog(
                    company=company,
                    staff=request_personnel,
                    action=action_description
                )
                log.save()
            obj_data = {'id': obj_id}  # ID bilgisini ekleyin
            for field in ModelClass._meta.fields:
                obj_data[field.name] = getattr(obj, field.name)
            context = {
                'form': form,
                'obj_data': obj_data,  # Güncellenmiş obje verileri
                'model_name': model,
                'fields': [field.name for field in ModelClass._meta.fields],
            }
            # Güncellenmiş verilerle liste şablonunu yeniden render et
            return render(request, 'tour/partials/mobile-card.html', context)
    else:
        form = FormClass(instance=obj)
        context = {
            'form': form,
            'obj_data': obj,  # Form başlangıcında objeyi gönder
            'model_name': model,
            'fields': [field.name for field in ModelClass._meta.fields],
        }

    return render(request, 'tour/partials/mobile-edit.html', context)



@login_required
def generic_mobile_cancel_view(request, model, obj_id):
    ModelClass = globals().get(model.capitalize())
    if ModelClass is None:
        return HttpResponseNotFound(f"{model.capitalize()} model not found.")
    # Model objesini getir veya 404 döndür
    obj = get_object_or_404(ModelClass, id=obj_id)
    # İlgili firma nesnesini al
    obj_data = {'id': obj_id}  # ID bilgisini ekleyin
    for field in ModelClass._meta.fields:
        obj_data[field.name] = getattr(obj, field.name)

    # Firma bilgilerini içeren bir template döndür
    return render(request, 'tour/partials/mobile-card.html', {'obj_data': obj_data, 'fields': [field.name for field in ModelClass._meta.fields], 'model_name': model,})

@login_required
def dark_mode(request):
    if request.method == "POST":
        personel = get_object_or_404(Personel, user=request.user)
        personel.dark_mode = not personel.dark_mode
        personel.save()

        # Kullanıcının geldiği URL'yi al
        referer_url = request.META.get('HTTP_REFERER')

        # Kullanıcının geldiği URL varsa oraya, yoksa başka bir URL'ye yönlendir
        return HttpResponseRedirect(referer_url if referer_url else '/fallback-url/')
    else:
        # POST olmayan isteklerde ana sayfaya yönlendirme
        return HttpResponseRedirect('/')
from django.core.paginator import Paginator
from django.db.models.functions import TruncDate

@login_required
def logs(request):
    if not can_access_personel2(request.user):
        return HttpResponseForbidden("Bu bölüme erişim yetkiniz bulunmamaktadır.")
    user = request.user
    requestPersonel = Personel.objects.get(user=request.user)
    sirket = requestPersonel.company
    staffs = Personel.objects.filter(company=sirket).order_by('user__first_name')
    logs = UserActivityLog.objects.filter(company=sirket).exclude(action="Giriş Yaptı.") \
        .annotate(date_only=TruncDate('timestamp')).order_by('-date_only', '-timestamp')
    page = Paginator(logs, 500)
    page_list = request.GET.get('page')
    page = page.get_page(page_list)
    context = {'page': page, 'title': 'Loglar', 'login': False, 'personel': None, 'staffs': staffs,}
    return render(request, 'tour/pages/logs.html', context)

@login_required
def log_staff(request, personel_id):
    staff = Personel.objects.get(id=personel_id)
    if not can_access_personel2(request.user):
        return HttpResponseForbidden("Bu bölüme erişim yetkiniz bulunmamaktadır.")
    user = request.user
    requestPersonel = Personel.objects.get(user=request.user)
    sirket = requestPersonel.company
    staffs = Personel.objects.filter(company=sirket).order_by('user__first_name')
    logs = UserActivityLog.objects.filter(company=sirket, staff=staff).exclude(action="Giriş Yaptı.") \
        .annotate(date_only=TruncDate('timestamp')).order_by('-date_only', '-timestamp')
    page = Paginator(logs, 500)
    page_list = request.GET.get('page')
    page = page.get_page(page_list)
    context = {'page': page, 'title': 'Loglar', 'login': False, 'personel': personel_id, 'staffs': staffs,}
    return render(request, 'tour/pages/logs.html', context)

@login_required
def login_logs(request):
    if not can_access_personel2(request.user):
        return HttpResponseForbidden("Bu bölüme erişim yetkiniz bulunmamaktadır.")
    user = request.user
    requestPersonel = Personel.objects.get(user=request.user)
    sirket = requestPersonel.company
    staffs = Personel.objects.filter(company=sirket).order_by('user__first_name')
    logs = UserActivityLog.objects.filter(company=sirket, action="Giriş Yaptı.") \
        .annotate(date_only=TruncDate('timestamp')).order_by('-date_only', '-timestamp')
    page = Paginator(logs, 500)
    page_list = request.GET.get('page')
    page = page.get_page(page_list)
    context = {'page': page, 'title': 'Giriş Logları', 'login': True, 'personel': None, 'staffs': staffs}
    return render(request, 'tour/pages/logs.html', context)

@login_required
def log_staff_login(request, personel_id):
    staff = Personel.objects.get(id=personel_id)
    if not can_access_personel2(request.user):
        return HttpResponseForbidden("Bu bölüme erişim yetkiniz bulunmamaktadır.")
    user = request.user
    requestPersonel = Personel.objects.get(user=request.user)
    sirket = requestPersonel.company
    staffs = Personel.objects.filter(company=sirket).order_by('user__first_name')
    logs = UserActivityLog.objects.filter(company=sirket, staff=staff, action="Giriş Yaptı.") \
        .annotate(date_only=TruncDate('timestamp')).order_by('-date_only', '-timestamp')
    page = Paginator(logs, 500)
    page_list = request.GET.get('page')
    page = page.get_page(page_list)
    context = {'page': page, 'title': 'Giriş Logları', 'login': True, 'personel': personel_id, 'staffs': staffs}
    return render(request, 'tour/pages/logs.html', context)


#####################################################################################################################################################
#####################################################################################################################################################
#####################################################################################################################################################
#####################################################################################################################################################
#####################################################################################################################################################


@login_required
def operation(request):
    requestPersonel = Personel.objects.get(user = request.user)
    sirket = requestPersonel.company

    context={

        'title': 'Operasyon Kur',
        'createtitle': 'Operasyon Oluştur',
        'createUrl' : 'create_operation',
        'deleteUrl' : 'delete_buyer',
        'form' : OperationForm(request=request),
        'formgun' : OperationdayForm(),
        'formitem' : OperationitemForm(request=request),


    }

    return render(request, 'tour/pages/operation.html', context)

@login_required
def create_operation(request):
    if request.method == "POST":
        requestPersonel = Personel.objects.get(user = request.user)
        sirket = requestPersonel.company
        form = OperationForm(request.POST or None, request=request)
        if form.is_valid():
            operasyon = form.save(commit=False)
            operasyon.selling_staff = requestPersonel
            operasyon.company = sirket
            operasyon.save()
            baslangic_tarihi = operasyon.start
            bitis_tarihi = operasyon.finish

            gun_sayisi = (bitis_tarihi - baslangic_tarihi).days
            for i in range(gun_sayisi + 1):
                gun = baslangic_tarihi + timedelta(days=i)
                operasyon_gun_form = OperationdayForm({'date': gun})
                if operasyon_gun_form.is_valid():
                    operasyon_gun = operasyon_gun_form.save(commit=False)
                    operasyon_gun.operation = operasyon
                    operasyon_gun.company = sirket
                    operasyon_gun.save()
            try:
                UserActivityLog.objects.create(staff=requestPersonel, company=sirket, action=f"Operasyon kaydı yapıldı. Operasyon ID: {operasyon.id} Operasyon Etiket: {operasyon.ticket}")
            except Personel.DoesNotExist:
                pass  # veya uygun bir hata mesajı göster
            notification_text = f"Yeni bir operasyon oluşturuldu: Operasyon ID: {operasyon.id}, Operasyon Etiket: {operasyon.ticket}"
            notification_title = "Tur Oluşturuldu. [Otomatik Bildirim]"
            sender = requestPersonel
            recipients_group = "Herkes"
            notification = Notification(company=sirket, sender=sender, recipients_group=recipients_group, title=notification_title, message=notification_text)
            notification.save()
            recipients = Personel.objects.filter(company=sirket, is_delete=False)
            for recipient in recipients:
                NotificationReceipt.objects.create(notification=notification, recipient=recipient)
            context = {
                'operasyon': operasyon,
                'days' : Operationday.objects.filter(company=sirket, operation=operasyon),
                'formitem' : OperationitemForm(request=request)
            }
            return render(request, 'tour/partials/operation-day.html', context)

    context = {'form': OperationForm()}
    return render(request, 'tour/pages/operation.html', context)

@login_required
def create_operation_item(request, day_id):
    day = get_object_or_404(Operationday, id=day_id)
    if request.method == "POST":
        requestPersonel = Personel.objects.get(user=request.user)
        sirket = requestPersonel.company

        formitem = OperationitemForm(request.POST, request=request)
        if formitem.is_valid():
            new = formitem.save(commit=False)
            new.company = sirket
            new.day = day

            try:
                UserActivityLog.objects.create(staff=requestPersonel, company=sirket, action=f"Operasyonİtem kaydı yapıldı. Operasyonitem ID: {new.id} Operasyon Etiket: {new.day.operation.ticket} İşlem Türü: {new.operation_type}")
            except Personel.DoesNotExist:
                pass

            new.save()
            formitem.save_m2m()

            return HttpResponse('Başarıyla Kaydedildi.')
        else:
            print('Form geçersiz. Hatalar: ' + str(formitem.errors))

    context = {
        'formitem': OperationitemForm(request=request),
        'day': day,
        'random_number': random.randint(1000, 9999)
    }
    return render(request, 'tour/partials/operation-item-form.html', context)

@login_required
def create_operation_item_add(request):
    context = {
        'formitem': OperationitemForm(request=request),
        'random_number': random.randint(1000, 9999)
    }
    return render(request, 'tour/partials/operation-item-form.html', context)
#####################################################################################################
#####################################################################################################
#####################################################################################################
#####################################################################################################


@login_required
def operation_list(request):
    requestPersonel = Personel.objects.get(user=request.user)
    sirket = requestPersonel.company
    today = datetime.today()

    buyer_companies = Buyercompany.objects.filter(company=sirket, is_delete=False).order_by('name')

    # Her bir buyer_company için ilgili operasyonları getirin
    for company in buyer_companies:
        company.started_operations = Operation.objects.filter(company=sirket, start__lte=today, finish__gte=today, is_delete=False, buyer_company=company).order_by('start')
        company.finished_operations = Operation.objects.filter(company=sirket, finish__lt=today, is_delete=False, buyer_company=company).order_by('-start')
        company.future_operations = Operation.objects.filter(company=sirket, start__gt=today, is_delete=False, buyer_company=company).order_by('start')

    return render(request, 'tour/pages/operation-list.html', {
        'buyer_companies': buyer_companies,
        'title': 'Operasyon',
        'createtitle': 'Operasyon Listesi',
    })
from django.db.models import Prefetch

@login_required
def operation_details(request, operation_id):
    # Retrieve the operation with related objects in a single query
    operation_day_prefetch = Prefetch('days', queryset=Operationday.objects.order_by('date').prefetch_related('items'))
    # Operation nesnesini ilgili nesnelerle birlikte tek bir sorguda alın
    operation = Operation.objects.filter(id=operation_id).prefetch_related(operation_day_prefetch).first()

    # Check if operation exists
    if not operation:
        # Handle the case where the operation doesn't exist, e.g., return a 404 response
        return HttpResponseNotFound('Operation not found')

    update_operation_costs(operation)

    context = {
        'operation': operation,
    }

    return render(request, 'tour/partials/operation-details.html', context)

@login_required
def update_operation(request, operation_id):
    requestPersonel = Personel.objects.get(user=request.user)
    sirket = requestPersonel.company
    operation = get_object_or_404(Operation, id=operation_id)
    follow_staff = operation.follow_staff
    buyer_company = operation.buyer_company
    start = operation.start
    finish = operation.finish
    ticket = operation.ticket
    passenger_info = operation.passenger_info
    usd_sales_price = operation.usd_sales_price
    tl_sales_price = operation.tl_sales_price
    eur_sales_price = operation.eur_sales_price
    rbm_sales_price = operation.rbm_sales_price
    number_passengers = operation.number_passengers
    payment_channel = operation.payment_channel
    payment_type = operation.payment_type
    sold = operation.sold
    if request.method == "POST":
        form = OperationForm(request.POST, instance=operation, request=request)
        if form.is_valid():
            updated_operation = form.save(commit=False)
            print(f"Eski: {tl_sales_price}, Yeni: {updated_operation.tl_sales_price}")
            if any([
                follow_staff != updated_operation.follow_staff,
                buyer_company != updated_operation.buyer_company,
                start != updated_operation.start,
                finish != updated_operation.finish,
                passenger_info != updated_operation.passenger_info,
                ticket != updated_operation.ticket,
                usd_sales_price != updated_operation.usd_sales_price,
                tl_sales_price != updated_operation.tl_sales_price,
                eur_sales_price != updated_operation.eur_sales_price,
                rbm_sales_price != updated_operation.rbm_sales_price,
                number_passengers != updated_operation.number_passengers,
                sold != updated_operation.sold,
                payment_channel != updated_operation.payment_channel,
                payment_type != updated_operation.payment_type
            ]):
                updated_operation.save()  # Formu kaydet

                # Bitiş tarihi değişikliğini kontrol et
                if updated_operation.finish > finish:
                    current_date = finish + timedelta(days=1)
                    while current_date <= updated_operation.finish:
                        Operationday.objects.get_or_create(
                            date=current_date,
                            operation=updated_operation,
                            defaults={'company': sirket}
                        )
                        current_date += timedelta(days=1)

                days_with_items = Operationday.objects.filter(operation=updated_operation).annotate(items_count=Count('items')).filter(items_count__gt=0)

                if updated_operation.finish < finish:
                    max_date = days_with_items.filter(date__gt=updated_operation.finish).aggregate(max_date=Max('date'))['max_date']
                    if max_date:
                        updated_operation.finish = max_date
                        updated_operation.save()

                    Operationday.objects.filter(operation=updated_operation, date__gt=updated_operation.finish).exclude(id__in=days_with_items.values('id')).delete()

                if updated_operation.start > start:
                    min_date = days_with_items.filter(date__lt=updated_operation.start).aggregate(min_date=Min('date'))['min_date']
                    if min_date:
                        updated_operation.start = min_date
                        updated_operation.save()

                    Operationday.objects.filter(operation=updated_operation, date__lt=updated_operation.start).exclude(id__in=days_with_items.values('id')).delete()

                if updated_operation.start < start:
                    current_date = updated_operation.start
                    while current_date < start:
                        Operationday.objects.get_or_create(
                            date=current_date,
                            operation=updated_operation,
                            defaults={'company': sirket}
                        )
                        current_date += timedelta(days=1)

                action = f"Operasyon Güncellendi. Operasyon ID : {updated_operation.id}"
                # Hangi alanların değiştiğini kontrol et
                if follow_staff != updated_operation.follow_staff:
                    action += f" Takip Eden Personel: {follow_staff}>{updated_operation.follow_staff}"
                if buyer_company != updated_operation.buyer_company:
                    action += f" Alıcı Şirket: {buyer_company}>{updated_operation.buyer_company}"
                if start != updated_operation.start:
                    action += f" Başlangıç Tarihi: {start}>{updated_operation.start}"
                if finish != updated_operation.finish:
                    action += f" Bitiş Tarihi: {finish}>{updated_operation.finish}"
                if passenger_info != updated_operation.passenger_info:
                    action += f" Yolcu Bilgisi: {passenger_info}>{updated_operation.passenger_info}"
                if ticket != updated_operation.ticket:
                    action += f" Etiket: {ticket}>{updated_operation.ticket}"
                if usd_sales_price != updated_operation.usd_sales_price:
                    action += f" USD Satış Fiyatı: {usd_sales_price}>{updated_operation.usd_sales_price}"
                if tl_sales_price != updated_operation.tl_sales_price:
                    action += f" TL Satış Fiyatı: {tl_sales_price}>{updated_operation.tl_sales_price}"
                if eur_sales_price != updated_operation.eur_sales_price:
                    action += f" EUR Satış Fiyatı: {eur_sales_price}>{updated_operation.eur_sales_price}"
                if rbm_sales_price != updated_operation.rbm_sales_price:
                    action += f" RBM Satış Fiyatı: {rbm_sales_price}>{updated_operation.rbm_sales_price}"
                if number_passengers != updated_operation.number_passengers:
                    action += f" Yolcu Sayısı: {number_passengers}>{updated_operation.number_passengers}"
                if payment_type != updated_operation.payment_type:
                    action += f" Ödeme Türü: {payment_type}>{updated_operation.payment_type}"
                if payment_channel != updated_operation.payment_channel:
                    action += f" Ödeme Kanalı: {payment_channel}>{updated_operation.payment_channel}"

                try:
                    UserActivityLog.objects.create(staff=requestPersonel, company=sirket, action=action)
                except Personel.DoesNotExist:
                    pass

                notification_text = f"Ticket: {updated_operation.ticket} Güncellendi. ID{updated_operation.id} {action}"
                notification_title = "Bir Operation Güncellendi. [Otomatik Bildirim]"
                sender = requestPersonel
                recipients_group = "Herkes"
                notification = Notification(company=sirket, sender=sender, recipients_group=recipients_group, title=notification_title, message=notification_text)
                notification.save()
                recipients = Personel.objects.filter(company=sirket, is_delete=False)
                for recipient in recipients:
                    NotificationReceipt.objects.create(notification=notification, recipient=recipient)

                update_operation_costs(updated_operation)
                messages.success(request, "Operasyon başarıyla güncellendi.")
                return HttpResponse('Başarı ile kaydedildi.')
            else:
                messages.info(request, "Hiçbir değişiklik yapılmadı.")
                return HttpResponse('Başarı ile kaydedildi.')
        else:
            print(f"Operation Hata2: {form.errors}")
            return HttpResponse('Form kaydedilemedi.')
    else:
        form = OperationForm(instance=operation, request=request)

    operation_day_forms = [
        (OperationdayForm(instance=day), [OperationitemForm(instance=item, request=request) for item in day.items.filter(is_delete=False)])
        for day in operation.days.filter(is_delete=False).order_by('date')
    ]


    return render(request, 'tour/pages/update_operation.html', {'operation_form': form, 'title': 'Güncelle', 'operation_day_forms': operation_day_forms, 'operation': operation})

def update_total_cost(operation_item, currency, price):
    if price:
        if price != 0 and price != Decimal('0.00'):
            if operation_item.tl_cost_price is None:
                operation_item.tl_cost_price = Decimal('0.0')
            if operation_item.usd_cost_price is None:
                operation_item.usd_cost_price = Decimal('0.0')
            if operation_item.eur_cost_price is None:
                operation_item.eur_cost_price = Decimal('0.0')
            if operation_item.rmb_cost_price is None:
                operation_item.rmb_cost_price = Decimal('0.0')

            if currency == "TL":
                operation_item.tl_cost_price += price
            elif currency == "USD":
                operation_item.usd_cost_price += price
            elif currency == "EUR":
                operation_item.eur_cost_price += price
            elif currency == "RMB":
                operation_item.rmb_cost_price += price
            operation_item.save()

def update_operation_costs(operation):


    # Tüm maliyetlerin toplamını sıfırla
    total_tl = total_usd = total_eur = total_rmb = 0

    # Operation'a bağlı tüm Operationitem'ları al ve topla
    for item in Operationitem.objects.filter(day__operation=operation):
        total_tl += item.tl_cost_price
        total_usd += item.usd_cost_price
        total_eur += item.eur_cost_price
        total_rmb += item.rmb_cost_price



    # Operation modelindeki maliyet alanlarını güncelle
    operation.tl_cost_price = total_tl
    operation.usd_cost_price = total_usd
    operation.eur_cost_price = total_eur
    operation.rbm_cost_price = total_rmb
    operation.save()

def add_operation_item(request, day_id):
    requestPersonel = Personel.objects.get(user=request.user)
    sirket = requestPersonel.company
    operation_day = get_object_or_404(Operationday, id=day_id)
    operation = operation_day.operation

    if request.method == "POST":
        bugun = datetime.now().date()
        yarin = bugun + timedelta(days=1)
        ertesigun = bugun + timedelta(days=2)
        formitem = OperationitemForm(request.POST, request=request)
        if formitem.is_valid():
            operation_item = formitem.save(commit=False)
            operation_item.day = operation_day
            operation_item.company = sirket
            operation_item.save()
            formitem.save_m2m()  # ManyToMany alanları kaydet
            operation_item.refresh_from_db()
            return HttpResponse('Başarıyla Güncellendi.')
        else:
            print('Form geçersiz. Hatalar: ' + str(formitem.errors))
    else:
        formitem = OperationitemForm(request=request)

    return render(request, 'tour/partials/add_operation_item.html', {'formitem': formitem, 'day_id': day_id, 'operation_day': operation_day, 'random_number': random.randint(1000, 9999)})

@login_required
def update_or_add_operation_item(request, day_id, item_id=None):
    requestPersonel = Personel.objects.get(user=request.user)
    sirket = requestPersonel.company
    operation_day = get_object_or_404(Operationday, id=day_id)
    operation = operation_day.operation
    operation_item = None if item_id is None else get_object_or_404(Operationitem, id=item_id)
    action = ""

    if operation_item and item_id:
        old_driver = operation_item.driver
        old_driver_phone = operation_item.driver_phone if operation_item.driver else None
        old_plaka = operation_item.plaka if operation_item.driver else None
        old_guide = operation_item.guide
        old_guide_phone = operation_item.guide.phone if operation_item.guide else None
        old_description = operation_item.description
        old_pick_time = operation_item.pick_time
        old_pick_location = operation_item.pick_location
        old_tour = operation_item.tour
        old_transfer = operation_item.transfer

    old_values = {}
    if operation_item and item_id:
        old_values = model_to_dict(operation_item, exclude=['id', 'day'])
        for field, value in old_values.items():
            if isinstance(getattr(operation_item, field), Manager):  # ManyToManyField'ları kontrol et
                old_values[field] = set(getattr(operation_item, field).all())
            elif hasattr(getattr(operation_item, field), '__str__'):
                old_values[field] = str(getattr(operation_item, field))
        action = f"Operasyon İçeriği Güncellendi. Operasyon Etiketi: {operation_item.day.operation.ticket}, Operasyon İtem ID: {operation_item.id}. Değişenler: "


    if request.method == "POST":
        bugun= datetime.now().date()
        yarin = bugun + timedelta(days=1)
        ertesigun = bugun + timedelta(days=2)
        formitem = OperationitemForm(request.POST, instance=operation_item, request=request)
        if formitem.is_valid():
            operation_item = formitem.save(commit=False)
            operation_item.day = operation_day
            operation_item.company = sirket
            vehicle_price = 0
            operation_item.tl_cost_price = 0
            operation_item.usd_cost_price = 0
            operation_item.eur_cost_price = 0
            operation_item.rmb_cost_price = 0
            if not operation_item.vehicle_price:
                operation_item.vehicle_price = 0

            operation_item.save()
            formitem.save_m2m()  # ManyToMany alanları kaydet
            operation_item.refresh_from_db()
            if operation_item.manuel_vehicle_price and operation_item.manuel_vehicle_price != 0:
                operation_item.vehicle_price = operation_item.manuel_vehicle_price
                vehicle_currency = operation_item.vehicle_currency
                vehicle_price = operation_item.vehicle_price
                if operation_item.vehicle_price is None:
                    operation_item.vehicle_price = 0

                operation_item.save()
                operation_item.refresh_from_db()
                update_total_cost(operation_item, vehicle_currency, vehicle_price)
            else:
                if operation_item.tour and operation_item.tour.route != "WALKING TUR IST":
                    if operation_item.supplier and operation_item.vehicle:
                        cost = Cost.objects.filter(supplier=operation_item.supplier, tour=operation_item.tour).first()
                        if cost:
                            if operation_item.vehicle == Vehicle.objects.get(id=38):
                                vehicle = "BINEK"
                                auot_vehicle_price = cost.car
                            if operation_item.vehicle == Vehicle.objects.get(id=39):
                                vehicle = "MINIVAN"
                                auot_vehicle_price = cost.minivan
                            if operation_item.vehicle == Vehicle.objects.get(id=40):
                                vehicle = "MINIBUS"
                                auot_vehicle_price = cost.minibus
                            if operation_item.vehicle == Vehicle.objects.get(id=41):
                                vehicle = "MIDIBUS"
                                auot_vehicle_price = cost.midibus
                            if operation_item.vehicle == Vehicle.objects.get(id=42):
                                vehicle = "OTOBUS"
                                auot_vehicle_price = cost.bus
                            operation_item.cost = cost
                            operation_item.auot_vehicle_price = auot_vehicle_price
                            operation_item.vehicle_price = auot_vehicle_price
                            operation_item.vehicle_currency = cost.currency
                            vehicle_currency = cost.currency
                            vehicle_price = operation_item.vehicle_price
                            update_total_cost(operation_item, vehicle_currency, vehicle_price)
                        else:
                            print('cost bulunamadı')
                if operation_item.transfer:
                    if operation_item.supplier and operation_item.vehicle:
                        cost = Cost.objects.filter(supplier=operation_item.supplier, transfer=operation_item.transfer).first()
                        if cost:
                            if operation_item.vehicle == Vehicle.objects.get(id=38):
                                vehicle = "BINEK"
                                auot_vehicle_price = cost.car
                            if operation_item.vehicle == Vehicle.objects.get(id=39):
                                vehicle = "MINIVAN"
                                auot_vehicle_price = cost.minivan
                            if operation_item.vehicle == Vehicle.objects.get(id=40):
                                vehicle = "MINIBUS"
                                auot_vehicle_price = cost.minibus
                            if operation_item.vehicle == Vehicle.objects.get(id=41):
                                vehicle = "MIDIBUS"
                                auot_vehicle_price = cost.midibus
                            if operation_item.vehicle == Vehicle.objects.get(id=42):
                                vehicle = "OTOBUS"
                                auot_vehicle_price = cost.bus
                            operation_item.cost = cost
                            operation_item.auot_vehicle_price = auot_vehicle_price
                            operation_item.vehicle_price = auot_vehicle_price
                            operation_item.vehicle_currency = cost.currency
                            vehicle_currency = cost.currency
                            vehicle_price = operation_item.vehicle_price
                            update_total_cost(operation_item, vehicle_currency, vehicle_price)
                        else:
                            print('cost bulunamadı')
            if operation_item.manuel_activity_price and operation_item.manuel_activity_price != 0:
                operation_item.activity_price = operation_item.manuel_activity_price
                activity_price = operation_item.activity_price
                activity_currency = operation_item.activity_currency
                operation_item.save()
                operation_item.refresh_from_db()
                update_total_cost(operation_item, activity_currency, activity_price)
            else:
                if operation_item.activity and operation_item.activity_supplier:
                    activity_cost = Activitycost.objects.filter(activity=operation_item.activity, supplier=operation_item.activity_supplier).first()
                    if activity_cost:
                        print(activity_cost)
                        operation_item.activity_cost = activity_cost
                        operation_item.auto_activity_price = activity_cost.price
                        operation_item.activity_currency = activity_cost.currency
                        operation_item.activity_price = operation_item.auto_activity_price
                        activity_price = activity_cost.price
                        activity_currency = activity_cost.currency
                        if operation_item.activity_payment == "Evet":
                            update_total_cost(operation_item, activity_currency, activity_price)
                else:
                    print('aktivite cost bulunamadı')

            museum_currency = operation_item.museum_currency
            museum_price = operation_item.museum_price
            if operation_item.museum_payment == "Evet":
                update_total_cost(operation_item, museum_currency, museum_price)
            hotel_currency = operation_item.hotel_currency
            hotel_price = operation_item.hotel_price

            guide_currency = operation_item.guide_currency
            guide_price = operation_item.guide_price
            other_currency = operation_item.other_currency
            other_price = operation_item.other_price
            update_total_cost(operation_item, other_currency, other_price)
            update_total_cost(operation_item, guide_currency, guide_price)
            if operation_item.hotel_payment == "Evet":
                update_total_cost(operation_item, hotel_currency, hotel_price)

            operation_item.save()
            operation_item.refresh_from_db()

            new_values = model_to_dict(operation_item, exclude=['id', 'day'])
            changes_detected = False  # Değişiklik algılanıp algılanmadığını izle
            for field, new_value in new_values.items():
                old_value = old_values.get(field)
                if old_value is None:  # Eğer old_value None ise
                    old_value = None
                if isinstance(getattr(operation_item, field), Manager):
                    new_value = set(getattr(operation_item, field).all())
                elif hasattr(getattr(operation_item, field), '__str__'):
                    new_value = str(getattr(operation_item, field))
                if new_value != old_value:
                    changes_detected = True
                    verbose_name = operation_item._meta.get_field(field).verbose_name
                    if isinstance(new_value, set) and old_value is not None:
                        removed = old_value - new_value if old_value is not None else set()
                        added = new_value - old_value if new_value is not None else set()
                        if removed:
                            action += f"Çıkarılan {verbose_name}: {', '.join(str(x) for x in removed)} "
                        if added:
                            action += f"Eklenen {verbose_name}: {', '.join(str(x) for x in added)} "
                    else:
                        action += f"{verbose_name} : {old_value} > {new_value} "

                # action uzunluğunu kontrol et ve gerekiyorsa kes



            if action and changes_detected:
                UserActivityLog.objects.create(staff=requestPersonel, company=sirket, action=action)

            if action != f"Operasyon İçeriği Güncellendi. Operasyon Etiketi: {operation_item.day.operation.ticket}, Operasyon İtem ID: {operation_item.id}. Değişenler: ":
                if operation_item.day.date == bugun or operation_item.day.date == yarin or operation_item.day.date == ertesigun:
                    notification_text = f"Ticket: {operation_item.day.operation.ticket} Gün: {operation_item.day.date} İşlem Türü: {operation_item.operation_type} ID: {operation_item.id} {action}"
                    notification_title = "3 gün içerisindeki bir iş değiştirildi. [Otomatik Bildirim]"
                    sender = requestPersonel
                    recipients_group = "Herkes"
                    notification = Notification(company=sirket, sender=sender, recipients_group=recipients_group, title=notification_title, message=notification_text)
                    notification.save()
                    recipients = Personel.objects.filter(company=sirket, is_delete=False)
                    for recipient in recipients:
                        NotificationReceipt.objects.create(notification=notification, recipient=recipient)

            # Şoför veya rehber değişikliği kontrol et
            if old_driver != operation_item.driver:
                new_driver = operation_item.driver
                new_driver_phone = operation_item.driver_phone
                if old_driver:
                    sms("driver", "cancelled", operation_item,old_driver, old_driver_phone)
                sms("driver", "assigned", operation_item, new_driver, new_driver_phone)
                if operation_item.guide:
                    new_guide = operation_item.guide.name
                    new_guide_phone = operation_item.guide.phone
                    sms("guide", "assigned", operation_item, new_guide, new_guide_phone)
            elif operation_item.driver == None and old_driver:
                sms("driver", "cancelled", operation_item, old_driver, old_driver_phone)

            if old_guide != operation_item.guide:
                if operation_item.guide:
                    new_guide = operation_item.guide.name
                    new_guide_phone = operation_item.guide.phone
                    print(new_guide, new_guide_phone)
                    if operation_item.driver:
                        sms("guide", "assigned", operation_item, new_guide, new_guide_phone)
                if old_guide:
                    sms("guide", "cancelled", operation_item, old_guide, old_guide_phone)
            elif operation_item.guide == None and old_guide:
                sms("guide", "cancelled", operation_item, old_guide, old_guide_phone)

            if old_description != operation_item.description:
                if operation_item.driver and operation_item.description:
                    new_driver = operation_item.driver
                    new_driver_phone = operation_item.driver_phone
                    sms("driver", "assigned", operation_item, new_driver, new_driver_phone)
                if operation_item.guide and operation_item.description:
                    new_guide = operation_item.guide.name
                    new_guide_phone = operation_item.guide.phone
                    sms("guide", "assigned", operation_item, new_guide, new_guide_phone)

            if old_pick_time != operation_item.pick_time:
                if operation_item.driver and operation_item.pick_time:
                    new_driver = operation_item.driver
                    new_driver_phone = operation_item.driver_phone
                    sms("driver", "assigned", operation_item, new_driver, new_driver_phone)
                if operation_item.guide and operation_item.pick_time:
                    new_guide = operation_item.guide.name
                    new_guide_phone = operation_item.guide.phone
                    sms("guide", "assigned", operation_item, new_guide, new_guide_phone)

            if old_pick_location != operation_item.pick_location:
                if operation_item.driver and operation_item.pick_location:
                    new_driver = operation_item.driver
                    new_driver_phone = operation_item.driver_phone
                    sms("driver", operation_item, "assigned", new_driver, new_driver_phone)
                if operation_item.guide and operation_item.pick_location:
                    new_guide = operation_item.guide.name
                    new_guide_phone = operation_item.guide.phone
                    sms("guide", "assigned", operation_item, new_guide, new_guide_phone)
            if old_tour:
                if old_tour != operation_item.tour:
                    if operation_item.driver and operation_item.tour:
                        new_driver = operation_item.driver
                        new_driver_phone = operation_item.driver_phone
                        sms("driver", "assigned", operation_item, new_driver, new_driver_phone)
                    if operation_item.guide and operation_item.tour:
                        new_guide = operation_item.guide.name
                        new_guide_phone = operation_item.guide.phone
                        sms("guide", "assigned", operation_item, new_guide, new_guide_phone)
                elif old_transfer != operation_item.transfer:
                    if operation_item.driver and operation_item.transfer:
                        new_driver = operation_item.driver
                        new_driver_phone = operation_item.driver_phone
                        sms("driver", "assigned", operation_item, new_driver, new_driver_phone)
                    if operation_item.guide and operation_item.transfer:
                        new_guide = operation_item.guide.name
                        new_guide_phone = operation_item.guide.phone
                        sms("guide", "assigned", operation_item, new_guide, new_guide_phone)


            return HttpResponse('Başarıyla Güncellendi.')
        else:
            print('Form geçersiz. Hatalar: ' + str(formitem.errors))
    else:
        formitem = OperationitemForm(request=request)

    return render(request, 'tour/partials/update_formitem_add.html', {'formitem': formitem, 'day_id': day_id, 'operation_day': operation_day, 'random_number': random.randint(1000, 9999)})

def sms(staff, action, item=None, old_value=None, phone=None):
    url = "http://soap.netgsm.com.tr:8080/Sms_webservis/SMS?wsdl"
    headers = {'content-type': 'text/xml'}
    mesaj = ""
    if item:
        operation_detail = item.tour if item.tour else item.transfer
    else:
        operation_detail = "Belirtilmemiş"

    if staff == "driver":
        if action == "assigned":
            mesaj = f"Sayın {old_value}, {item.day.date.day}.{item.day.date.month}.{item.day.date.year} tarihinde {item.operation_type} tanımlanmıştır. Açıklaması: {item.description}, Güzergah: {operation_detail}, Alış Saati: {item.pick_time}, Alış Yeri: {item.pick_location}, Bırakış Yeri: {item.release_location}"
        elif action == "cancelled":
            mesaj = f"Sayın {old_value}, {item.day.date.day}.{item.day.date.month}.{item.day.date.year} tarihinde {item.operation_type} iptal edilmiştir."

    elif staff == "guide":
        if action == "assigned":
            mesaj = f"Sayın {old_value}, {item.day.date.day}.{item.day.date.month}.{item.day.date.year} tarihinde {item.operation_type} tanımlanmıştır. Açıklaması: {item.description}, Güzergah: {operation_detail}, Alış Saati: {item.pick_time}, Alış Yeri: {item.pick_location}, Bırakış Yeri: {item.release_location}. Şoför {item.driver}, Tel: {item.driver_phone}, Plaka {item.plaka}"
        elif action == "cancelled":
            mesaj = f"Sayın {old_value}, {item.day.date.day}.{item.day.date.month}.{item.day.date.year} tarihinde {item.operation_type} iptal edilmiştir."

    elif staff == "destek":
        if action == "olusturuldu":
             mesaj = f"Yeni destek kaydı oluşturuldu."

    body = f"""<?xml version="1.0"?>
        <SOAP-ENV:Envelope xmlns:SOAP-ENV="http://schemas.xmlsoap.org/soap/envelope/"
                    xmlns:xsd="http://www.w3.org/2001/XMLSchema"
                    xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">
            <SOAP-ENV:Body>
                <ns3:smsGonder1NV2 xmlns:ns3="http://sms/">
                    <username>8503081334</username>
                    <password>6D18AD8</password>
                    <header>MNC GROUP</header>
                    <msg>{mesaj}</msg>
                    <gsm>{phone}</gsm>
                    <encoding>TR</encoding>
                    <filter>0</filter>
                    <startdate></startdate>
                    <stopdate></stopdate>
                </ns3:smsGonder1NV2>
            </SOAP-ENV:Body>
        </SOAP-ENV:Envelope>"""

    response = requests.post(url, data=body, headers=headers)
    print("response")
    print("response.status_code:", response.status_code)
    print("response.text:", response.text)


def delete_operation(request, operation_id):
    requestPersonel = Personel.objects.get(user=request.user)
    sirket = requestPersonel.company
    if request.method == "DELETE":
        operation = get_object_or_404(Operation, id=operation_id)
        ticket = operation.ticket
        try:
            UserActivityLog.objects.create(staff=requestPersonel, company=sirket, action=f"Operasyon kaydı silindi. Operasyon ID: {operation.id} Operasyon Etiket: {operation.ticket}")
        except Personel.DoesNotExist:
            pass  # veya uygun bir hata mesajı göster
        operation.is_delete = True
        base_ticket = f"{ticket}Silindi"
        operation.ticket = base_ticket
        suffix = 1
        while Operation.objects.filter(ticket=operation.ticket).exists():
            operation.ticket = f"{base_ticket}{suffix}"
            suffix += 1
        operation.save()
         # İlgili gün ve itemları sil
        Operationday.objects.filter(operation=operation).update(is_delete=True)
        Operationitem.objects.filter(day__operation=operation).update(is_delete=True)
        return HttpResponse('Başarıyla Silindi.')  # Boş bir yanıt döndür

@login_required
def delete_operationitem(request, operationitem_id):
    requestPersonel = Personel.objects.get(user=request.user)
    sirket = requestPersonel.company
    if request.method == "DELETE":
        operationitem = get_object_or_404(Operationitem, id=operationitem_id)
        try:
            UserActivityLog.objects.create(staff=requestPersonel, company=sirket, action=f"Operasyon İtem kaydı silindi. Operasyon İtem ID: {operationitem.id} Operasyon Gün: {operationitem.day.date} Operasyon Etiket: {operationitem.day.operation.ticket}")
        except Personel.DoesNotExist:
            pass  # veya uygun bir hata mesajı göster
        operationitem.delete()
        return HttpResponse('')  # Boş bir yanıt döndür


def check_ticket(request):
    ticket = request.GET.get('ticket', '')
    is_taken = Operation.objects.filter(ticket=ticket).exists()
    return JsonResponse({'is_taken': is_taken})




@login_required
def index(request):
    requestPersonel = Personel.objects.get(user=request.user)
    sirket = requestPersonel.company

    if request.method == "POST":
        date = request.POST.get('date_job')
        finishdate = request.POST.get('date_finish_job') if request.POST.get('date_finish_job') != "" else None

        if finishdate:
            week_jobs = Operationitem.objects.filter(company=sirket, is_delete=False, day__date__range=(date, finishdate))
            date_parts = date.split('-')
            date_finish_parts = finishdate.split('-')

            reversed_date = '.'.join(date_parts[::-1])
            reversed_date2 = '.'.join(date_finish_parts[::-1])
            weektitle = f"{reversed_date} - {reversed_date2} Tarihli İşler"
        else:
            week_jobs = Operationitem.objects.filter(company=sirket, day__date=date, is_delete=False)
            reversed_date = '.'.join(date.split('-')[::-1])
            weektitle = f"{reversed_date} Tarihli İşler"
    else:
        week_jobs = None
        weektitle = "Tarih Giriniz"

    today = datetime.today()
    tomorrow = today + timedelta(days=1)
    totomorrow = today + timedelta(days=2)

    today_jobs = Operationitem.objects.filter(company=sirket, day__date=today, is_delete=False).order_by('pick_time')
    tomorrow_jobs = Operationitem.objects.filter(company=sirket, day__date=tomorrow, is_delete=False).order_by('pick_time')
    totomorrow_jobs = Operationitem.objects.filter(company=sirket, day__date=totomorrow, is_delete=False).order_by('pick_time')

    def set_row_class(jobs):
        for job in jobs:
            if job.operation_type in ['Tur', 'TurTransfer', 'TransferTur']:
                if job.vehicle and job.supplier and job.pick_time:
                    if job.hotel_payment == 'Yes':
                        if job.hotel and job.hotel_price:
                            if job.museum_payment == 'Yes':
                                if job.new_museum and job.museum_price:
                                    if job.activity_payment == 'Yes':
                                        if job.activity and job.activity_price and job.activity_supplier:
                                            if job.guide_var == 'Yes':
                                                if job.guide and job.guide_price != 0:
                                                    job.row_class = ""
                                                else:
                                                    job.row_class = "red_tr"
                                            else:
                                                job.row_class = ""
                                        else:
                                            job.row_class = "red_tr"
                                    else:
                                        job.row_class = ""
                                else:
                                    job.row_class = "red_tr"
                            else:
                                job.row_class = ""
                        else:
                            job.row_class = "red_tr"
                    else:
                        job.row_class = ""
                else:
                    job.row_class = "red_tr"
            elif job.operation_type == 'Transfer':
                if job.vehicle and job.supplier and job.pick_time:
                    if job.hotel_payment == 'Yes':
                        if job.hotel and job.room_type and job.hotel_price:
                            if job.museum_payment == 'Yes':
                                if job.new_museum and job.museum_price:
                                    if job.activity_payment == 'Yes':
                                        if job.activity and job.activity_price and job.activity_supplier:
                                            if job.guide_var == 'Yes':
                                                if job.guide and job.guide_price != 0:
                                                    job.row_class = ""
                                                else:
                                                    job.row_class = "red_tr"
                                            else:
                                                job.row_class = ""
                                        else:
                                            job.row_class = "red_tr"
                                    else:
                                        job.row_class = ""
                                else:
                                    job.row_class = "red_tr"
                            else:
                                job.row_class = ""
                        else:
                            job.row_class = "red_tr"
                    else:
                        job.row_class = ""
                else:
                    job.row_class = "red_tr"
            elif job.operation_type == 'Aktivite':
                if job.activity_payment == 'Yes' and job.activity and job.pick_time and job.activity_price and job.activity_supplier:
                    job.row_class = ""
                else:
                    job.row_class = "red_tr"
            elif job.operation_type == 'Müze':
                if job.museum_payment == 'Yes' and job.new_museum and job.pick_time and job.museum_price:
                    job.row_class = ""
                else:
                    job.row_class = "red_tr"
            elif job.operation_type == 'Rehber':
                if job.guide_var == 'Yes' and job.guide and job.guide_price and job.guide_currency:
                    job.row_class = ""
                else:
                    job.row_class = "red_tr"
            else:
                job.row_class = ""

    set_row_class(today_jobs)
    set_row_class(tomorrow_jobs)
    set_row_class(totomorrow_jobs)
    if week_jobs:
        set_row_class(week_jobs)

    context = {
        'today_jobs': today_jobs,
        'tomorrow_jobs': tomorrow_jobs,
        'totomorrow_jobs': totomorrow_jobs,
        'week_jobs': week_jobs,
        'todaytitle': today,
        'tomorrowtitle': tomorrow,
        'totomorrowtitle': totomorrow,
        'weektitle': weektitle,
        'title': 'İşler',
        'test': 'test%20Mesajıdır',
    }

    return render(request, 'tour/pages/index.html', context)

@login_required
def filtre(request):
    requestPersonel = Personel.objects.get(user=request.user)
    today = datetime.today()
    tomorrow = today + timedelta(days=1)
    totomorrow = today + timedelta(days=2)
    sirket = requestPersonel.company
    if request.method == "POST":
        filtre = request.POST.get('filtre')
        today_jobs = Operationitem.objects.filter(operation_type=filtre, company=sirket, day__date=today).order_by('pick_time')
        tomorrow_jobs = Operationitem.objects.filter(operation_type=filtre, company=sirket, day__date=tomorrow).order_by('pick_time')
        totomorrow_jobs = Operationitem.objects.filter(operation_type=filtre,company=sirket, day__date=totomorrow).order_by('pick_time')
    else:
        today_jobs = Operationitem.objects.filter(company=sirket, day__date=today).order_by('pick_time')
        tomorrow_jobs = Operationitem.objects.filter(company=sirket, day__date=tomorrow).order_by('pick_time')
        totomorrow_jobs = Operationitem.objects.filter(company=sirket, day__date=totomorrow).order_by('pick_time')

    # Python'da sıralama yapın

    # Context
    context = {
        'today_jobs': today_jobs,
        'tomorrow_jobs': tomorrow_jobs,
        'totomorrow_jobs': totomorrow_jobs,
        'todaytitle': today,
        'tomorrowtitle': tomorrow,
        'totomorrowtitle': totomorrow,
        'title': 'İşler'
    }

    return render(request, 'tour/pages/index.html', context)

@login_required
def indexim(request):
    today = datetime.today()
    tomorrow = today + timedelta(days=1)
    totomorrow = today + timedelta(days=2)
    requestPersonel = Personel.objects.get(user=request.user)
    sirket = requestPersonel.company
    exchange_rates()

    if request.method == "POST":
        date = request.POST.get('date_job')
        if request.POST.get('date_finish_job') != "":
            finishdate = request.POST.get('date_finish_job')
        else:
            finishdate=None
        if finishdate != None:
            week_jobs = Operationitem.objects.filter(company=sirket, is_delete=False, day__operation__follow_staff=requestPersonel, day__date__range=(date, finishdate))
            date_parts = date.split('-')
            date_finish_parts = finishdate.split('-')

            # Listenin sırasını tersine çevir
            reversed_date_parts = date_parts[::-1]
            reversed_date_parts2 = date_finish_parts[::-1]

            # Ters çevrilmiş listeyi yazdır
            reversed_date = '.'.join(reversed_date_parts)
            reversed_date2 = '.'.join(reversed_date_parts2)
            weektitle = f"{reversed_date} - {reversed_date2} Tarihli İşler"
        else:
            week_jobs = Operationitem.objects.filter(company=sirket, day__date=date, is_delete=False, day__operation__follow_staff=requestPersonel)
            date_parts = date.split('-')

            # Listenin sırasını tersine çevir
            reversed_date_parts = date_parts[::-1]

            # Ters çevrilmiş listeyi yazdır
            reversed_date = '.'.join(reversed_date_parts)
            weektitle = f"{reversed_date} Tarihli İşler"


    else:
        week_jobs = None
        weektitle = "Tarih Giriniz"
    # Dates


    # Queries
    today_jobs = Operationitem.objects.filter(company=sirket, day__date=today, day__operation__follow_staff=requestPersonel, is_delete=False).order_by('pick_time')
    tomorrow_jobs = Operationitem.objects.filter(company=sirket, day__date=tomorrow, day__operation__follow_staff=requestPersonel, is_delete=False).order_by('pick_time')
    totomorrow_jobs = Operationitem.objects.filter(company=sirket, day__date=totomorrow, day__operation__follow_staff=requestPersonel, is_delete=False).order_by('pick_time')
    # Python'da sıralama yapın

    def set_row_class(jobs):
        for job in jobs:
            if job.operation_type in ['Tur', 'TurTransfer', 'TransferTur']:
                if job.vehicle and job.supplier and job.pick_time:
                    if job.hotel_payment == 'Yes':
                        if job.hotel and job.room_type and job.hotel_price:
                            if job.museum_payment == 'Yes':
                                if job.new_museum and job.museum_price:
                                    if job.activity_payment == 'Yes':
                                        if job.activity and job.activity_price and job.activity_supplier:
                                            if job.guide_var == 'Yes':
                                                if job.guide and job.guide_price != 0:
                                                    job.row_class = ""
                                                else:
                                                    job.row_class = "red_tr"
                                            else:
                                                job.row_class = ""
                                        else:
                                            job.row_class = "red_tr"
                                    else:
                                        job.row_class = ""
                                else:
                                    job.row_class = "red_tr"
                            else:
                                job.row_class = ""
                        else:
                            job.row_class = "red_tr"
                    else:
                        job.row_class = ""
                else:
                    job.row_class = "red_tr"
            elif job.operation_type == 'Transfer':
                if job.vehicle and job.supplier and job.pick_time:
                    if job.hotel_payment == 'Yes':
                        if job.hotel and job.room_type and job.hotel_price:
                            if job.museum_payment == 'Yes':
                                if job.new_museum and job.museum_price:
                                    if job.activity_payment == 'Yes':
                                        if job.activity and job.activity_price and job.activity_supplier:
                                            if job.guide_var == 'Yes':
                                                if job.guide and job.guide_price != 0:
                                                    job.row_class = ""
                                                else:
                                                    job.row_class = "red_tr"
                                            else:
                                                job.row_class = ""
                                        else:
                                            job.row_class = "red_tr"
                                    else:
                                        job.row_class = ""
                                else:
                                    job.row_class = "red_tr"
                            else:
                                job.row_class = ""
                        else:
                            job.row_class = "red_tr"
                    else:
                        job.row_class = ""
                else:
                    job.row_class = "red_tr"
            elif job.operation_type == 'Aktivite':
                if job.activity_payment == 'Yes' and job.activity and job.pick_time and job.activity_price and job.activity_supplier:
                    job.row_class = ""
                else:
                    job.row_class = "red_tr"
            elif job.operation_type == 'Müze':
                if job.museum_payment == 'Yes' and job.new_museum and job.pick_time and job.museum_price:
                    job.row_class = ""
                else:
                    job.row_class = "red_tr"
            elif job.operation_type == 'Rehber':
                if job.guide_var == 'Yes' and job.guide and job.guide_price and job.guide_currency:
                    job.row_class = ""
                else:
                    job.row_class = "red_tr"
            else:
                job.row_class = ""

    set_row_class(today_jobs)
    set_row_class(tomorrow_jobs)
    set_row_class(totomorrow_jobs)
    if week_jobs:
        set_row_class(week_jobs)
    # Context
    context = {
        'today_jobs': today_jobs,
        'tomorrow_jobs': tomorrow_jobs,
        'totomorrow_jobs': totomorrow_jobs,
        'week_jobs': week_jobs,
        'todaytitle': today,
        'tomorrowtitle': tomorrow,
        'totomorrowtitle': totomorrow,
        'weektitle': weektitle,
        'title': 'İşler'
    }

    return render(request, 'tour/pages/index.html', context)
from django.utils.timezone import make_aware

def create_notification(request):
    if not can_access_personel2(request.user):
        return HttpResponseForbidden("Bu bölüme erişim yetkiniz bulunmamaktadır.")
    requestPersonel = Personel.objects.get(user=request.user)
    today = date.today()
    start_of_today = make_aware(datetime.combine(today, time.min))
    end_of_today = make_aware(datetime.combine(today, time.max))

    print(today)
    sirket = requestPersonel.company
    if request.method == 'POST':

        form = NotificationForm(request.POST)
        if form.is_valid():
            notification = form.save(commit=False)
            notification.sender = requestPersonel
            notification.company = sirket
            notification.save()
            job_title = form.cleaned_data['recipients_group']
            if notification.sender.job == "Sistem Geliştiricisi":
                if job_title != "Herkes":
                    recipients = Personel.objects.filter(job=job_title, is_delete=False)
                else:
                    recipients = Personel.objects.filter(is_delete=False)
                for recipient in recipients:
                    NotificationReceipt.objects.create(notification=notification, recipient=recipient)

            else:
                if job_title != "Herkes":
                    recipients = Personel.objects.filter(company=sirket, job=job_title, is_delete=False)
                else:
                    recipients = Personel.objects.filter(company=sirket, is_delete=False)
                for recipient in recipients:
                    NotificationReceipt.objects.create(notification=notification, recipient=recipient)

            messages.success(request, 'Notification successfully created.')
            return redirect('notification_create') # Başarıyla oluşturulduktan sonra yönlendirilecek yer
    else:
        form = NotificationForm()
        if requestPersonel.job == "Sistem Geliştiricisi":
            my_notifications = Notification.objects.filter(
                sender=requestPersonel,
                timestamp__range=(start_of_today, end_of_today)
            ).prefetch_related(
                Prefetch('receipts', queryset=NotificationReceipt.objects.all())
            ).order_by('-timestamp')  # Azalan sıralama için '-timestamp'
        else:
            my_notifications = Notification.objects.filter(
                company=sirket,
                sender=requestPersonel,
                timestamp__range=(start_of_today, end_of_today)
            ).prefetch_related(
                Prefetch('receipts', queryset=NotificationReceipt.objects.filter(recipient__company=sirket))
            ).order_by('-timestamp')  # Azalan sıralama için '-timestamp'


    return render(request, 'tour/pages/create_notification.html', {'form': form, 'my_notifications' : my_notifications, 'title' : 'Bildirim Oluştur'})

@csrf_exempt  # Eğer CSRF token kullanmıyorsanız
def notifications_api(request):
    if request.method == 'GET':
        try:
            personel = Personel.objects.get(user=request.user)
            company = personel.company
        except AttributeError:
            return JsonResponse({'error': 'Personel object not found for this user'}, status=404)




        today = datetime.now().date()
        yesterday = today - timedelta(days=1)

        notifications = NotificationReceipt.objects.filter(
            Q(recipient=personel, notification__company=company) |
            Q(notification__sender__job="Sistem Geliştiricisi", recipient=personel),
            notification__timestamp__date__in=[yesterday, today],
            read_at__isnull=True
        ).order_by('notification__timestamp')

        notifications_data = [
            {
                'id': n.id,
                'title': n.notification.title,
                'message': n.notification.message,
                'timestamp': n.notification.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
                'sender': f"{n.notification.sender.user.first_name} {n.notification.sender.user.last_name} - {n.notification.sender.job}" if n.notification.sender else None
            } for n in notifications
        ]

        return JsonResponse({'notifications': notifications_data})

    elif request.method == 'POST':
        try:
            data = json.loads(request.body)
            notification_id = data['notification_id']
            personel_id = data.get('personel_id')
            if not personel_id:
                return JsonResponse({'error': 'Personel ID is required'}, status=400)
            personel = Personel.objects.get(id=personel_id)

            notification = NotificationReceipt.objects.get(id=notification_id, recipient=personel)
            if notification.read_at:
                return JsonResponse({'status': 'error', 'message': 'Notification already marked as read'})
            notification.read_at = timezone.now()
            notification.save()
            return JsonResponse({'status': 'success', 'message': 'Notification marked as read'})
        except NotificationReceipt.DoesNotExist:
            return JsonResponse({'error': 'Notification not found or not available for this user'}, status=404)
        except Personel.DoesNotExist:
            return JsonResponse({'error': 'Personel not found'}, status=404)
        except KeyError:
            return JsonResponse({'error': 'Incorrect data format'}, status=400)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

def mark_notification_as_read(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        notification_id = data.get('notification_id')
        personel = Personel.objects.get(user = request.user)  # Kullanıcı oturum bilgisini kullanarak alınır

        try:
            receipt = NotificationReceipt.objects.get(notification__id=notification_id, recipient=personel)
            receipt.read_at = date.now()
            receipt.save()
            return JsonResponse({'status': 'success', 'message': 'Notification marked as read.'})
        except NotificationReceipt.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Notification not found.'}, status=404)




def template_create(request, operation_id):
    requestPersonel = Personel.objects.get(user= request.user)
    sirket = requestPersonel.company
    operation = get_object_or_404(Operation, id=operation_id)
    ticket_temp = operation.ticket
    if request.method == "POST":
        template_name = request.POST.get('template_name')

        # Operasyona ait günleri ve ilişkili itemları çek
        operation_days_with_items = Operationday.objects.prefetch_related(
            'operationitem_set'
        ).filter(operation=operation)

        # OperationTemplate oluştur
        operation_template = OperationTemplate.objects.create(
            company=operation.company,
            name=template_name,
            buyer_company=operation.buyer_company,
            day_numbers=operation_days_with_items.count(),
            tl_sales_price=operation.tl_sales_price,
            usd_sales_price=operation.usd_sales_price,
            eur_sales_price=operation.eur_sales_price,
            rbm_sales_price=operation.rbm_sales_price
        )

        # Her bir gün ve günün itemları için döngü
        for index, day in enumerate(operation_days_with_items, start=1):
            day_template = OperationdayTemplate.objects.create(
                operation=operation_template,
                company=operation.company,
                day_number=index
            )

            # Günün itemlarını template'e kopyala
            for item in day.operationitem_set.all():
                item_template = OperationitemTemplate.objects.create(
                    company=operation.company,
                    operation=operation_template,
                    day=day_template,
                    operation_type=item.operation_type,
                    pick_time=item.pick_time,
                    release_time=item.release_time,
                    tour=item.tour,
                    transfer=item.transfer,
                    vehicle=item.vehicle,
                    hotel=item.hotel,
                    room_type=item.room_type,
                    hotel_payment=item.hotel_payment,
                    activity=item.activity,
                    activity_payment=item.activity_payment,
                    description=item.description
                )
                # ManyToManyField ilişkileri için müzeleri ayarla
                item_template.new_museum.set(item.new_museum.all())

        try:
            UserActivityLog.objects.create(staff=requestPersonel, company=sirket, action=f"Operasyon Şablonu Oluşturuldu. Operasyon ID: {operation_id} Operasyon Etiket: {ticket_temp} Şablon Adı: {template_name}")
        except Personel.DoesNotExist:
            pass  # veya uygun bir hata mesajı göster

        return HttpResponse('Operasyon şablonu başarıyla oluşturuldu.')
    return HttpResponse('Operasyon şablonu oluşturulamadı.')




def template_list(request):
    request_personel = Personel.objects.get(user=request.user)
    sirket = request_personel.company
    if request.method == "POST":
        start_date_str = request.POST.get('start_date')
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        template_id = request.POST.get('template_id')
        template = get_object_or_404(OperationTemplate, id=template_id)

        short_name = template.buyer_company.short_name.upper() if template.buyer_company else 'NN'
        formatted_date = start_date.strftime('%d%m%y')
        ticket_base = f"{short_name}{formatted_date}"

        for i in range(100):
            new_ticket = f"{ticket_base}{i:02}" if i != 0 else ticket_base
            if not Operation.objects.filter(ticket=new_ticket).exists():
                operation = Operation.objects.create(
                    ticket=new_ticket, company=sirket, buyer_company=template.buyer_company, selling_staff=request_personel, follow_staff=request_personel,
                    tl_sales_price=template.tl_sales_price, usd_sales_price=template.usd_sales_price,
                    eur_sales_price=template.eur_sales_price, rbm_sales_price=template.rbm_sales_price,
                    start=start_date, finish=start_date + timedelta(days=template.day_numbers - 1)
                )
                break
        else:
            return HttpResponse('Uygun bir ticket numarası bulunamadı.', status=400)

        # Tüm item şablonlarını bir kez çek
        operation_items_templates = OperationitemTemplate.objects.filter(operation=template, company=sirket)
        # Her gün için Operationday ve Operationitem'ları oluştur
        for day_offset in range(template.day_numbers):
            operation_date = start_date + timedelta(days=day_offset)
            operasyon_gun = Operationday.objects.create(
                date=operation_date, operation=operation, company=sirket)
            # Bu gün için uygun olan itemları bul ve oluştur
            day_set = OperationdayTemplate.objects.filter(operation=template, day_number=day_offset + 1, company=sirket)
            for day in day_set:
                day_items = operation_items_templates.filter(day=day)
                for item_template in day_items:
                    print(item_template)
                    operation_item = Operationitem.objects.create(
                        operation_type=item_template.operation_type, day=operasyon_gun, company=sirket,
                        pick_time=item_template.pick_time, release_time=item_template.release_time,
                        tour=item_template.tour, transfer=item_template.transfer, vehicle=item_template.vehicle,
                        hotel=item_template.hotel, room_type=item_template.room_type,
                        hotel_payment=item_template.hotel_payment, activity=item_template.activity,
                        activity_payment=item_template.activity_payment, museum_payment=item_template.museum_payment,
                        description=item_template.description
                    )
                    operation_item.new_museum.set(item_template.new_museum.all())

        try:
            UserActivityLog.objects.create(staff=request_personel, company=sirket, action=f"Şablondan Operasyon Oluşturuldu. Operasyon ID: {operation.id} Operasyon Etiket: {new_ticket} Şablon Adı: {template.name}")
        except Personel.DoesNotExist:
            pass

        return render(request, 'tour/pages/template_list.html')
    else:
        templates = OperationTemplate.objects.filter(company=sirket)
        context = {'templates': templates}
        return render(request, 'tour/pages/template_list.html', context)


def cari(request):
    requestPersonel = Personel.objects.get(user=request.user)
    sirket = requestPersonel.company
    suppliers = Supplier.objects.filter(company=sirket, is_delete=False).order_by('name')

    context={
        'suppliers' : suppliers,
        'activity' : False,
        'buyer_gelir' : False,
    }
    return render(request, 'tour/pages/cari.html', context)

def buyer(request):
    requestPersonel = Personel.objects.get(user=request.user)
    sirket = requestPersonel.company
    buyers = Buyercompany.objects.filter(company=sirket, is_delete=False).order_by('name')

    context={
        'suppliers' : buyers,
        'buyer_gelir' : True,
        'activity' : False
    }
    return render(request, 'tour/pages/cari.html', context)

def buyer_category(request, tedarikci_id, month=None):
    toplam_rmb_bakiye = 0
    toplam_usd_bakiye = 0
    toplam_eur_bakiye = 0
    toplam_tl_bakiye = 0
    toplam_bakiye = 0
    requestPersonel = Personel.objects.get(user=request.user)
    sirket = requestPersonel.company
    tedarikci = get_object_or_404(Buyercompany, id=tedarikci_id)
    month = datetime.now().month
    items = Operation.objects.filter(company=sirket, buyer_company=tedarikci, start__month=month, is_delete=False)
    if items:
        for item in items:
            satis = 0
            exchange = ExchangeRate.objects.filter(created_at__gte=item.start).order_by('created_at').first()
            if exchange:
                eur_sales_price = Decimal(item.eur_sales_price or 0)
                tl_sales_price = Decimal(item.tl_sales_price or 0)
                rbm_sales_price = Decimal(item.rbm_sales_price or 0)
                usd_sales_price = Decimal(item.usd_sales_price or 0)

                usd_to_eur = Decimal(exchange.usd_to_eur)
                usd_to_try = Decimal(exchange.usd_to_try)
                usd_to_rmb = Decimal(exchange.usd_to_rmb)

                satis += (eur_sales_price / usd_to_eur) + (tl_sales_price / usd_to_try) + (rbm_sales_price / usd_to_rmb) + usd_sales_price
            else:
                exchange = ExchangeRate.objects.filter(created_at__lte=item.start).order_by('created_at').first()
                if exchange:
                    eur_sales_price = Decimal(item.eur_sales_price or 0)
                    tl_sales_price = Decimal(item.tl_sales_price or 0)
                    rbm_sales_price = Decimal(item.rbm_sales_price or 0)
                    usd_sales_price = Decimal(item.usd_sales_price or 0)

                    usd_to_eur = Decimal(exchange.usd_to_eur)
                    usd_to_try = Decimal(exchange.usd_to_try)
                    usd_to_rmb = Decimal(exchange.usd_to_rmb)

                    satis += (eur_sales_price / usd_to_eur) + (tl_sales_price / usd_to_try) + (rbm_sales_price / usd_to_rmb) + usd_sales_price
            toplam_tl_bakiye += item.tl_sales_price
            toplam_usd_bakiye += item.usd_sales_price
            toplam_eur_bakiye += item.eur_sales_price
            toplam_rmb_bakiye += item.rbm_sales_price
            item.total_sales_price = satis
            item.save()
            toplam_bakiye += item.total_sales_price

    if request.method == "POST":
        month_input = request.POST.get('month', '')
        month = datetime.now().month if not month_input else datetime.strptime(month_input, '%Y-%m').month
        tedarikci = get_object_or_404(Buyercompany, id=tedarikci_id)

        items = Operation.objects.filter(company=sirket, buyer_company=tedarikci, start__month=month, is_delete=False)

        toplam_rmb_bakiye = 0
        toplam_usd_bakiye = 0
        toplam_eur_bakiye = 0
        toplam_tl_bakiye = 0
        toplam_bakiye = 0

        for item in items:
            satis = 0
            exchange = ExchangeRate.objects.filter(created_at__gte=item.start).order_by('created_at').first()
            if exchange:
                eur_sales_price = Decimal(item.eur_sales_price or 0)
                tl_sales_price = Decimal(item.tl_sales_price or 0)
                rbm_sales_price = Decimal(item.rbm_sales_price or 0)
                usd_sales_price = Decimal(item.usd_sales_price or 0)

                usd_to_eur = Decimal(exchange.usd_to_eur)
                usd_to_try = Decimal(exchange.usd_to_try)
                usd_to_rmb = Decimal(exchange.usd_to_rmb)

                satis += (eur_sales_price / usd_to_eur) + (tl_sales_price / usd_to_try) + (rbm_sales_price / usd_to_rmb) + usd_sales_price
            else:
                exchange = ExchangeRate.objects.filter(created_at__lte=item.start).order_by('created_at').first()
                if exchange:
                    eur_sales_price = Decimal(item.eur_sales_price or 0)
                    tl_sales_price = Decimal(item.tl_sales_price or 0)
                    rbm_sales_price = Decimal(item.rbm_sales_price or 0)
                    usd_sales_price = Decimal(item.usd_sales_price or 0)

                    usd_to_eur = Decimal(exchange.usd_to_eur)
                    usd_to_try = Decimal(exchange.usd_to_try)
                    usd_to_rmb = Decimal(exchange.usd_to_rmb)

                    satis += (eur_sales_price / usd_to_eur) + (tl_sales_price / usd_to_try) + (rbm_sales_price / usd_to_rmb) + usd_sales_price
            toplam_tl_bakiye += item.tl_sales_price
            toplam_usd_bakiye += item.usd_sales_price
            toplam_eur_bakiye += item.eur_sales_price
            toplam_rmb_bakiye += item.rbm_sales_price
            item.total_sales_price = satis
            item.save()
            toplam_bakiye += item.total_sales_price

        context = {
            'tedarikci_id': tedarikci_id,
            'tedarikci' : tedarikci,
            'items': items,
            'month': month,
            'toplam_bakiye': toplam_bakiye,
            'toplam_rmb_bakiye': toplam_rmb_bakiye,
            'toplam_usd_bakiye': toplam_usd_bakiye,
            'toplam_eur_bakiye': toplam_eur_bakiye,
            'toplam_tl_bakiye': toplam_tl_bakiye,
            'activity': False,
            'buyer_gelir': True,
        }

        if not items.exists():  # Öğeler yoksa mesajı güncelle
            context['message'] = 'Uygun iş bulunamadı'
        return render(request, 'tour/partials/tedarikci_isleri.html', context)

    context = {
        'tedarikci_id': tedarikci_id,
        'tedarikci' : tedarikci,
        'items': items,
        'month' : month,
        'toplam_bakiye' : toplam_bakiye,
        'toplam_rmb_bakiye' : toplam_rmb_bakiye,
        'toplam_usd_bakiye' : toplam_usd_bakiye,
        'toplam_eur_bakiye' : toplam_eur_bakiye,
        'toplam_tl_bakiye' : toplam_tl_bakiye,
        'activity' : False,
        'buyer_gelir' : True,
    }

    return render(request, 'tour/partials/cost_categories.html', context)

def cari_category(request, tedarikci_id, month=None, field=None):
    toplam_bakiye = 0
    field = unquote(field) if field else field
    requestPersonel = Personel.objects.get(user=request.user)
    sirket = requestPersonel.company
    tedarikci = get_object_or_404(Supplier, id=tedarikci_id)
    month = datetime.now().month
    cost = Cost.objects.filter(company=sirket, supplier=tedarikci)
    items = Operationitem.objects.filter(company=sirket, supplier=tedarikci, day__date__month=month)
    if items:
        for item in items:
            if item.vehicle_price:
                toplam_bakiye += item.vehicle_price
    if request.method == "POST":
        month_input = request.POST.get('month', '')
        field = request.POST.get('field', field)
        month = datetime.now().month if not month_input else datetime.strptime(month_input, '%Y-%m').month

        if field == "tour":
            items = Operationitem.objects.filter(company=sirket, supplier=tedarikci, day__date__month=month, tour__isnull=False)
        elif field == "transfer":
            items = Operationitem.objects.filter(company=sirket, supplier=tedarikci, day__date__month=month, transfer__isnull=False)
        else:

            items = Operationitem.objects.filter(company=sirket, supplier=tedarikci, day__date__month=month, tour__route=field)
            if not items.exists():
                items = Operationitem.objects.filter(company=sirket, supplier=tedarikci, day__date__month=month, transfer__route=field)

        context = {
            'categories': cost,
            'tedarikci_id': tedarikci_id,
            'items': items,
            'month' : month,
            'toplam_bakiye' : toplam_bakiye,
            'activity' : False,
            'buyer_gelir' : False,
        }

        if not items.exists():  # Öğeler yoksa mesajı güncelle
            context['message'] = 'Uygun iş bulunamadı'
        return render(request, 'tour/partials/tedarikci_isleri.html', context)

    context = {
        'categories': cost,
        'tedarikci_id': tedarikci_id,
        'items': items,
        'month' : month,
        'toplam_bakiye' : toplam_bakiye,
    }

    return render(request, 'tour/partials/cost_categories.html', context)

def check_cost_duplicate(request):
    try:
        requestPersonel = Personel.objects.get(user=request.user)
    except Personel.DoesNotExist:
        return JsonResponse({'error': 'Personel not found'}, status=404)

    sirket = requestPersonel.company
    if not sirket:
        return JsonResponse({'error': 'Şirket not found'}, status=404)

    tour = request.GET.get('tour')
    transfer = request.GET.get('transfer')
    supplier = request.GET.get('supplier')

    # Ensure that tour, transfer, and supplier are integers if they exist
    try:
        filter_args = {'tour': int(tour)} if tour else {'transfer': int(transfer)}
        filter_args['supplier'] = int(supplier) if supplier else None
    except ValueError:
        return JsonResponse({'error': 'Invalid parameters'}, status=400)

    filter_args['company'] = sirket

    exists = Cost.objects.filter(**filter_args).exists()
    return JsonResponse({'exists': exists})

def check_hotel_duplicate(request):
    requestPersonel = Personel.objects.get(user=request.user)
    sirket=requestPersonel.company
    name = request.GET.get('name')
    city = request.GET.get('city')

    # name ve city'nin boş olmamasını sağlayarak gereksiz sorguları önleyin
    if name and city:
        exists = Hotel.objects.filter(name=name, city=city, company=sirket).exists()
        return JsonResponse({'exists': exists})
    else:
        # Eğer name veya city boşsa, geçersiz parametre hatası dön
        return JsonResponse({'error': 'Missing parameters'}, status=400)

def check_activity_duplicate(request):
    requestPersonel = Personel.objects.get(user=request.user)
    sirket=requestPersonel.company
    supplier = request.GET.get('supplier')
    activity = request.GET.get('activity')

    # name ve city'nin boş olmamasını sağlayarak gereksiz sorguları önleyin
    if supplier and activity:
        exists = Activitycost.objects.filter(supplier=supplier, activity=activity, company=sirket).exists()
        return JsonResponse({'exists': exists})
    else:
        # Eğer name veya city boşsa, geçersiz parametre hatası dön
        return JsonResponse({'error': 'Missing parameters'}, status=400)

def check_transfer_duplicate(request):
    requestPersonel = Personel.objects.get(user=request.user)
    sirket=requestPersonel.company
    route = request.GET.get('route')

    # name ve city'nin boş olmamasını sağlayarak gereksiz sorguları önleyin
    if route:
        exists = Transfer.objects.filter(route=route, company=sirket).exists()
        return JsonResponse({'exists': exists})
    else:
        # Eğer name veya city boşsa, geçersiz parametre hatası dön
        return JsonResponse({'error': 'Missing parameters'}, status=400)

def check_tour_duplicate(request):
    requestPersonel = Personel.objects.get(user=request.user)
    sirket=requestPersonel.company
    route = request.GET.get('route')

    # name ve city'nin boş olmamasını sağlayarak gereksiz sorguları önleyin
    if route:
        exists = Tour.objects.filter(route=route, company=sirket).exists()
        return JsonResponse({'exists': exists})
    else:
        # Eğer name veya city boşsa, geçersiz parametre hatası dön
        return JsonResponse({'error': 'Missing parameters'}, status=400)


def activity_cari(request):
    requestPersonel = Personel.objects.get(user=request.user)
    sirket = requestPersonel.company
    suppliers = Activitysupplier.objects.filter(company=sirket)

    context={
        'suppliers' : suppliers,
        'activity' : True
    }
    return render(request, 'tour/pages/cari.html', context)

def activity_cari_category(request, tedarikci_id, month=None, field=None):
    field = unquote(field) if field else field
    requestPersonel = Personel.objects.get(user=request.user)
    sirket = requestPersonel.company
    tedarikci = get_object_or_404(Activitysupplier, id=tedarikci_id)
    if not month:
        month = datetime.now().month
    cost = Activitycost.objects.filter(company=sirket, supplier=tedarikci)
    items = Operationitem.objects.filter(company=sirket, activity_supplier=tedarikci, day__date__month=month)
    if request.method == "POST":
        if month:
            month_input = request.POST.get('month', '')
            field = request.POST.get('field', field)  # POST'tan field değeri al, yoksa önceden belirlenen değeri kullan
            month = datetime.now().month if not month_input else datetime.strptime(month_input, '%Y-%m').month
        if field:
            items = Operationitem.objects.filter(company=sirket, activity_supplier=tedarikci, day__date__month=month, activity__name=field)
        else:
            items = Operationitem.objects.filter(company=sirket, activity_supplier=tedarikci, day__date__month=month)

        context = {
            'activity' : True,
            'categories': cost,
            'tedarikci_id': tedarikci_id,
            'items': items,
            'month' : month
        }

        if not items.exists():  # Öğeler yoksa mesajı güncelle
            context['message'] = 'Uygun iş bulunamadı'
        return render(request, 'tour/partials/tedarikci_isleri.html', context)

    context = {
        'activity' : True,
        'categories': cost,
        'tedarikci_id': tedarikci_id,
        'items': items,
        'month' : month
    }

    return render(request, 'tour/partials/cost_categories.html', context)




def support_ticket_create(request):
    try:
        request_personel = Personel.objects.get(user=request.user)
        # Diğer işlemler...
    except Personel.DoesNotExist:
        raise PermissionDenied("Personel bulunamadı.")
    sirket = request_personel.company

    if request.method == 'POST':
        form = SupportTicketForm(request.POST)
        if form.is_valid():
            support_ticket = form.save(commit=False)
            support_ticket.user = request_personel
            support_ticket.company = sirket
            support_ticket.status = "open"
            support_ticket.save()
            sms("destek", "olusturuldu", None, "Nurhak", "05054471953")
            return redirect('support_ticket_create')
    else:
        form = SupportTicketForm()

    if request_personel.job == "Sistem Geliştiricisi":
        support_tickets = SupportTicket.objects.all().order_by("-created_at")
    else:
        support_tickets = SupportTicket.objects.filter(user=request_personel).order_by("-created_at")

    return render(request, 'tour/pages/support_ticket_form.html', {
        'form': form,
        'support_tickets': support_tickets,
        'title': 'Support'
    })

def support_cevap(request, sup_id):
    sup = get_object_or_404(SupportTicket, id=sup_id)
    request_personel = Personel.objects.get(user=request.user)
    sirket = request_personel.company

    if request.method == 'POST':
        form = SupportTicketCevapForm(request.POST, instance=sup)
        if form.is_valid():
            support_ticket = form.save()
            return redirect('support_ticket_create')
    else:
        form = SupportTicketCevapForm(instance=sup)

    if request_personel.job == "Sistem Geliştiricisi":
        support_tickets = SupportTicket.objects.all().order_by("-created_at")
    else:
        support_tickets = SupportTicket.objects.filter(user=request_personel).order_by("-created_at")

    return render(request, 'tour/pages/support_ticket_form.html', {
        'form': form,
        'support_tickets': support_tickets,
        'title': 'Support'
    })
def get_suppliers(request, item_id):
    item = get_object_or_404(Operationitem, id=item_id)

    if item.tour:
        costs = Cost.objects.filter(tour=item.tour)
    elif item.transfer:
        costs = Cost.objects.filter(transfer=item.transfer)
    else:
        costs = Cost.objects.none()

    suppliers = costs.values('supplier_id', 'supplier__name').distinct()
    supplier_list = [{'id': supplier['supplier_id'], 'name': supplier['supplier__name']} for supplier in suppliers]

    return JsonResponse(supplier_list, safe=False)

def ortak_cost(request):
    # Kullanıcının personel ve şirket bilgilerini al
    request_personel = Personel.objects.get(user=request.user)
    sirket = request_personel.company

    # Şirkete ait olan tüm Operationitem kayıtlarını filtrele
    items = Operationitem.objects.filter(company=sirket)

    processed_items = 0  # İşlenen item sayısını takip etmek için

    for item in items:
        if item.tour and item.tour.id in [4, 5, 6]:
            tour = item.tour
            date = item.day.date
            plaka = item.plaka
            ucret = item.vehicle_price

            # Aynı şirkete, tura, tarihe ve plakaya sahip ve henüz işlenmemiş Operation kayıtlarını bul
            xitems = Operationitem.objects.filter(
                company=sirket,
                tour=tour,
                day__date=date,
                plaka=plaka,
                is_processed=False
            )

            if xitems:
                pax = 0
                for x in xitems:
                    pax += x.museum_person

                if pax > 0:
                    k_ucret = ucret / pax
                    for x in xitems:
                        total = k_ucret * x.museum_person
                        x.vehicle_price = total
                        x.is_processed = True  # İşlemi yapıldığını işaretle
                        x.save()
                        processed_items += 1  # İşlenen item sayısını artır
                        print(f"Processed Operation: {x.id} with total vehicle price: {total}")

    context = {}
    return redirect("cari")
import re
def smsgonder(request):
    request_personel = Personel.objects.get(user=request.user)
    sirket = request_personel.company
    url = "http://soap.netgsm.com.tr:8080/Sms_webservis/SMS?wsdl"
    headers = {'content-type': 'text/xml'}
    mesaj = ""

    if request.method == 'POST':
        form = SmsgonderForm(request.POST, request=request)
        if form.is_valid():
            smsgonder = form.save(commit=False)
            smsgonder.company = sirket
            smsgonder.user = request_personel
            smsgonder.save()
            form.save_m2m()  # Many-to-many ilişkileri kaydetmek için save_m2m() metodunu çağırın

            mesaj = f"{smsgonder.message}\n{request_personel.user.first_name} {request_personel.user.last_name} - {request_personel.job}"
            for alici in smsgonder.staff.all():
                if alici.phone:
                    phone = alici.phone.strip()
                    if phone.startswith('('):
                        match = re.match(r'\((\d{3})\)\s*(\d{3})-(\d{4})', phone)
                        if match:
                            formatted_phone = match.group(1) + match.group(2) + match.group(3)
                        else:
                            continue
                    elif phone.startswith('5'):
                        formatted_phone = '0' + phone
                    else:
                        formatted_phone = phone

                    body = f"""<?xml version="1.0"?>
                        <SOAP-ENV:Envelope xmlns:SOAP-ENV="http://schemas.xmlsoap.org/soap/envelope/"
                                    xmlns:xsd="http://www.w3.org/2001/XMLSchema"
                                    xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">
                            <SOAP-ENV:Body>
                                <ns3:smsGonder1NV2 xmlns:ns3="http://sms/">
                                    <username>8503081334</username>
                                    <password>6D18AD8</password>
                                    <header>MNC GROUP</header>
                                    <msg>{mesaj}</msg>
                                    <gsm>{formatted_phone}</gsm>
                                    <encoding>TR</encoding>
                                    <filter>0</filter>
                                    <startdate></startdate>
                                    <stopdate></stopdate>
                                </ns3:smsGonder1NV2>
                            </SOAP-ENV:Body>
                        </SOAP-ENV:Envelope>"""
                    response = requests.post(url, data=body, headers=headers)

            return redirect('smsgonder')  # Smsgonder detay sayfasına yönlendirme
    else:
        form = SmsgonderForm(request=request)
        if request_personel.job == "Sistem Geliştiricisi":
            smsgonder = Smsgonder.objects.all()
        else:
            smsgonder = Smsgonder.objects.filter(company=sirket, user=request_personel)

    context = {
        'form': form,
        'smsgonder': smsgonder,
        'title': 'SMS Gönder'
    }
    return render(request, 'tour/pages/smsgonder.html', context)


def fetch_personnel_list(request):
    request_personel = Personel.objects.get(user=request.user)
    sirket = request_personel.company
    personnel = Personel.objects.filter(company=sirket).exclude(id=request_personel.id).values('id', 'user__first_name')
    return JsonResponse(list(personnel), safe=False)

def fetch_chat_messages(request, personnel_id):
    try:
        user1 = request.user
        user2 = User.objects.get(personel__id=personnel_id)
        room = ChatRoom.get_or_create_room(user1, user2)
        messages = Message.objects.filter(room=room).values('user__id', 'user__first_name', 'content', 'timestamp')
        return JsonResponse(list(messages), safe=False)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
def send_message(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            personnel_id = data.get('personnel_id')
            content = data.get('content')

            # Authenticated user
            user1 = request.user
            user2 = User.objects.get(personel__id=personnel_id)

            # Get or create chat room between the two users
            room = ChatRoom.get_or_create_room(user1, user2)

            # Create a new message
            message = Message.objects.create(user=user1, room=room, content=content)

            # Return the user first name and message content
            response_data = {
                'user': {
                    'id': user1.id,
                    'first_name': user1.first_name
                },
                'content': message.content
            }

            return JsonResponse(response_data)

        except Personel.DoesNotExist:
            return JsonResponse({'error': 'Personnel does not exist.'}, status=404)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    return JsonResponse({'error': 'Invalid request method.'}, status=405)


def gelir(request, month=None):
    request_personel = Personel.objects.get(user=request.user)
    sirket = request_personel.company
    if not month:
        month_sec = datetime.now().month
    else:
        month_sec = int(month)

    odemeler = Cari.objects.filter(company=sirket, create_date__month=month_sec, is_delete=False).order_by("-create_date")
    context={
        'title' : 'Gelir Gider',
        'odemeler'  : odemeler
    }

    return render(request, 'tour/pages/gelir.html', context)

def grafik(request):
    request_personel = Personel.objects.get(user=request.user)
    sirket = request_personel.company
    current_year = datetime.now().year
    monthly_data = {
        'aylar': [],
        'satis': [],
        'maliyet': [],
        'kar': [],
    }

    for month in range(1, 13):
        maliyet = Decimal('0.0')
        satis = Decimal('0.0')
        operations = Operation.objects.filter(finish__year=current_year, finish__month=month, company=sirket, is_delete=False)

        for operation in operations:
            update_operation_costs(operation)
            start = operation.start
            exchange = ExchangeRate.objects.filter(created_at__gte=start).order_by('created_at').first()
            if exchange:
                eur_cost_price = Decimal(operation.eur_cost_price or 0)
                tl_cost_price = Decimal(operation.tl_cost_price or 0)
                rbm_cost_price = Decimal(operation.rbm_cost_price or 0)
                usd_cost_price = Decimal(operation.usd_cost_price or 0)

                eur_sales_price = Decimal(operation.eur_sales_price or 0)
                tl_sales_price = Decimal(operation.tl_sales_price or 0)
                rbm_sales_price = Decimal(operation.rbm_sales_price or 0)
                usd_sales_price = Decimal(operation.usd_sales_price or 0)

                usd_to_eur = Decimal(exchange.usd_to_eur)
                usd_to_try = Decimal(exchange.usd_to_try)
                usd_to_rmb = Decimal(exchange.usd_to_rmb)

                maliyet += (eur_cost_price / usd_to_eur) + (tl_cost_price / usd_to_try) + (rbm_cost_price / usd_to_rmb) + usd_cost_price
                satis += (eur_sales_price / usd_to_eur) + (tl_sales_price / usd_to_try) + (rbm_sales_price / usd_to_rmb) + usd_sales_price
            else:
                exchange = ExchangeRate.objects.filter(created_at__lte=start).order_by('created_at').first()
                if exchange:
                    eur_cost_price = Decimal(operation.eur_cost_price or 0)
                    tl_cost_price = Decimal(operation.tl_cost_price or 0)
                    rbm_cost_price = Decimal(operation.rbm_cost_price or 0)
                    usd_cost_price = Decimal(operation.usd_cost_price or 0)

                    eur_sales_price = Decimal(operation.eur_sales_price or 0)
                    tl_sales_price = Decimal(operation.tl_sales_price or 0)
                    rbm_sales_price = Decimal(operation.rbm_sales_price or 0)
                    usd_sales_price = Decimal(operation.usd_sales_price or 0)

                    usd_to_eur = Decimal(exchange.usd_to_eur)
                    usd_to_try = Decimal(exchange.usd_to_try)
                    usd_to_rmb = Decimal(exchange.usd_to_rmb)

                    maliyet += (eur_cost_price / usd_to_eur) + (tl_cost_price / usd_to_try) + (rbm_cost_price / usd_to_rmb) + usd_cost_price
                    satis += (eur_sales_price / usd_to_eur) + (tl_sales_price / usd_to_try) + (rbm_sales_price / usd_to_rmb) + usd_sales_price

        kar = satis - maliyet
        monthly_data['aylar'].append(datetime(current_year, month, 1).strftime('%B'))
        monthly_data['satis'].append(float(satis))
        monthly_data['maliyet'].append(float(maliyet))
        monthly_data['kar'].append(float(kar))

    context = {
        'title' : 'Grafik',
        'monthly_data': monthly_data,
    }
    return render(request, 'tour/pages/grafik.html', context)



def dashboard(request, month_sec=None):
    request_personel = Personel.objects.get(user=request.user)  # Assuming you have imported Personel model
    sirket = request_personel.company
    current_year = datetime.now().year
    today = datetime.now().date()  # Get today's date

    if not month_sec:
        month_sec = datetime.now().month
    else:
        month_sec = int(month_sec)  # Ensure month is an integer

    monthly_data = {
        'aylar': [],
        'satis': [],
        'maliyet': [],
        'kar': [],
        'marj': [],
        'oran': [],
        'tamam': [],
        'devam': [],
        'gelecek': [],
    }

    for month in range(1, 13):
        maliyet = Decimal('0.0')
        satis = Decimal('0.0')
        operations = Operation.objects.filter(finish__year=current_year, finish__month=month, company=sirket, is_delete=False)
        tamam = 0
        devam = 0
        gelecek = 0
        tamam_is = 0
        devam_is = 0
        gelecek_is = 0
        for operation in operations:
            update_operation_costs(operation)
            start = operation.start
            finish = operation.finish

            if start > today:
                gelecek += 1
            elif finish > today and start < today:
                devam += 1
            elif today > finish:
                tamam += 1

            exchange = ExchangeRate.objects.filter(created_at__gte=start).order_by('created_at').first()
            if not exchange:
                exchange = ExchangeRate.objects.filter(created_at__lte=start).order_by('-created_at').first()

            if exchange:
                eur_cost_price = Decimal(operation.eur_cost_price or 0)
                tl_cost_price = Decimal(operation.tl_cost_price or 0)
                rbm_cost_price = Decimal(operation.rbm_cost_price or 0)
                usd_cost_price = Decimal(operation.usd_cost_price or 0)

                eur_sales_price = Decimal(operation.eur_sales_price or 0)
                tl_sales_price = Decimal(operation.tl_sales_price or 0)
                rbm_sales_price = Decimal(operation.rbm_sales_price or 0)
                usd_sales_price = Decimal(operation.usd_sales_price or 0)

                usd_to_eur = Decimal(exchange.usd_to_eur)
                usd_to_try = Decimal(exchange.usd_to_try)
                usd_to_rmb = Decimal(exchange.usd_to_rmb)

                try:
                    maliyet += (eur_cost_price / usd_to_eur) + (tl_cost_price / usd_to_try) + (rbm_cost_price / usd_to_rmb) + usd_cost_price
                except ZeroDivisionError:
                    pass  # Handle appropriate error handling
                try:
                    satis += (eur_sales_price / usd_to_eur) + (tl_sales_price / usd_to_try) + (rbm_sales_price / usd_to_rmb) + usd_sales_price
                except ZeroDivisionError:
                    pass  # Handle appropriate error handling

        kar = satis - maliyet
        marj = (Decimal('0.0') if satis == 0 else Decimal(kar) / Decimal(satis) * Decimal(100.00))
        oran = (Decimal('0.0') if maliyet == 0 else Decimal(kar) / Decimal(maliyet) * Decimal(100.00))
        monthly_data['aylar'].append(datetime(current_year, month, 1).strftime('%B'))
        monthly_data['satis'].append(float(satis))
        monthly_data['maliyet'].append(float(maliyet))
        monthly_data['kar'].append(float(kar))
        monthly_data['marj'].append(float(marj))
        monthly_data['oran'].append(float(oran))
        monthly_data['tamam'].append(tamam)
        monthly_data['devam'].append(devam)
        monthly_data['gelecek'].append(gelecek)

    if 1 <= month_sec <= 12:
        month_marj = monthly_data['marj'][month_sec - 1]
        month_oran = monthly_data['oran'][month_sec - 1]
        gelir = monthly_data['satis'][month_sec - 1]
        gider = monthly_data['maliyet'][month_sec - 1]
        kazanc = monthly_data['kar'][month_sec - 1]
        tamam_is = monthly_data['tamam'][month_sec - 1]
        devam_is = monthly_data['devam'][month_sec - 1]
        gelecek_is = monthly_data['gelecek'][month_sec - 1]
    else:
        month_marj = None
        month_oran = None
        gelir = None
        gider = None
        kazanc = None
        tamam_is = 0
        devam_is = 0
        gelecek_is = 0
    items = Operation.objects.filter(finish__year=current_year, finish__month=month_sec, company=sirket, is_delete=False)
    context = {
        'title': 'Grafik',
        'monthly_data': monthly_data,
        'month_marj': month_marj,
        'month_oran': month_oran,
        'gelir': gelir,
        'gider': gider,
        'kazanc': kazanc,
        'tamam_is' : tamam_is,
        'devam_is' : devam_is,
        'gelecek_is' : gelecek_is,
        'operations' : items,
        'var' : False,
    }

    return render(request, 'tour/pages/dashboard.html', context)


def hata_bildir(request):
    url = "http://soap.netgsm.com.tr:8080/Sms_webservis/SMS?wsdl"
    headers = {'content-type': 'text/xml'}
    personel = Personel.objects.get(user = request.user)
    company = personel.company
    if request.method == 'POST':
        item_id = request.POST.get('item')
        message = request.POST.get('message')
        item = get_object_or_404(Operationitem, id=item_id)
        mesaj = ""
        if item:
            mesaj = f"{item.day.operation.ticket} grup kodlu turun {item.day.date.day}.{item.day.date.month} tarihinde hata tespit edildi. \n Hata: {message} \n {personel.user.first_name} {personel.user.last_name} {personel.job}"
            alici = item.day.operation.follow_staff
            if alici.phone:
                phone = alici.phone.strip()
                if phone.startswith('('):
                    match = re.match(r'\((\d{3})\)\s*(\d{3})-(\d{4})', phone)
                    if match:
                        formatted_phone = match.group(1) + match.group(2) + match.group(3)
                    else:
                        pass
                elif phone.startswith('5'):
                    formatted_phone = '0' + phone
                else:
                    formatted_phone = phone

                body = f"""<?xml version="1.0"?>
                    <SOAP-ENV:Envelope xmlns:SOAP-ENV="http://schemas.xmlsoap.org/soap/envelope/"
                                xmlns:xsd="http://www.w3.org/2001/XMLSchema"
                                xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">
                        <SOAP-ENV:Body>
                            <ns3:smsGonder1NV2 xmlns:ns3="http://sms/">
                                <username>8503081334</username>
                                <password>6D18AD8</password>
                                <header>MNC GROUP</header>
                                <msg>{mesaj}</msg>
                                <gsm>{formatted_phone}</gsm>
                                <encoding>TR</encoding>
                                <filter>0</filter>
                                <startdate></startdate>
                                <stopdate></stopdate>
                            </ns3:smsGonder1NV2>
                        </SOAP-ENV:Body>
                    </SOAP-ENV:Envelope>"""
                response = requests.post(url, data=body, headers=headers)
                log = UserActivityLog(
                    company=company,
                    staff=personel,
                    action=f"Sms Gönderildi. SMS: {mesaj}"
                )
                log.save()
                return redirect(index)
            else:
                return redirect(indexim)

    return redirect(index)



def odeme_create(request):
    personel = Personel.objects.get(user=request.user)
    company = personel.company
    if request.method == 'POST':
        form = CariForm(request.POST, request.FILES, request=request)
        if form.is_valid():
            odeme = form.save(commit=False)  # commit=False kullanarak formu kaydetmeden önce instance'ı döndürüyoruz
            odeme.company = company  # company alanını ayarlıyoruz
            odeme.created_staff = personel  # created_staff alanını ayarlıyoruz
            odeme.save()  # şimdi instance'ı kaydediyoruz
            return redirect('gelir')
    else:
        form = CariForm(request=request)
    return render(request, 'tour/pages/odeme.html', {'form': form})

def odeme_update(request, pk):
    personel = Personel.objects.get(user=request.user)
    cari = get_object_or_404(Cari, pk=pk)
    if request.method == 'POST':
        form = CariForm(request.POST, request.FILES, instance=cari, request=request)
        if form.is_valid():
            form.save()
            return redirect('gelir')
    else:
        form = CariForm(instance=cari, request=request)
    return render(request, 'cari_form.html', {'form': form})

def odeme_delete(request, pk):
    cari = get_object_or_404(Cari, pk=pk)
    cari.is_delete = True
    cari.save()
    return redirect('gelir')
